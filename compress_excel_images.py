#!/usr/bin/env python3
"""Excel Image Compressor

Compress embedded images in Excel reports using smart strategies:
  - Downsample: cap longest edge to shrink oversized microscope images
  - Large JPEG (>200KB) → q=85 re-encode (visually lossless, 50-80% smaller)
  - Small JPEG (≤200KB) → keep as-is (avoid re-compression bloat)
  - PNG (has alpha) → flatten to white + JPEG q=90
  - PNG (no alpha) → convert to JPEG q=90

Usage:
    python compress_excel_images.py report.xlsx                     # outputs report_compressed.xlsx
    python compress_excel_images.py report.xlsx -o out.xlsx         # custom output path
    python compress_excel_images.py report.xlsx --max-side 2048     # downsample (recommended)
    python compress_excel_images.py report.xlsx --max-side 2048 --quality 80
    python compress_excel_images.py report.xlsx --dry-run           # stats only, no output
"""

import argparse
import io
import os
import sys
import warnings

# ── Tunable defaults ──
JPEG_QUALITY_DEFAULT = 90       # default (PNG→JPEG, small JPEGs)
JPEG_QUALITY_LARGE = 85         # large JPEG re-encode
JPEG_LARGE_THRESHOLD_KB = 200   # trigger re-encode above this size


def compress_workbook(src_path, dst_path, quality_large=JPEG_QUALITY_LARGE,
                      quality_default=JPEG_QUALITY_DEFAULT,
                      threshold_kb=JPEG_LARGE_THRESHOLD_KB,
                      max_side=None, dry_run=False):
    """Process all embedded images in an Excel file. Returns stats dict.

    Args:
        max_side: max pixel length of the longest edge. Images exceeding
                  this are proportionally downsampled. None = no limit.
    """

    from PIL import Image as PILImage
    PILImage.MAX_IMAGE_PIXELS = None  # allow >178M pixel images

    import openpyxl
    from openpyxl.drawing.image import Image as XLImage

    warnings.filterwarnings('ignore')

    print(f"Loading: {src_path}")
    wb = openpyxl.load_workbook(src_path)

    stats = {
        'png_to_jpeg':    {'count': 0, 'before': 0, 'after': 0},
        'jpeg_reencode':  {'count': 0, 'before': 0, 'after': 0, 'skipped': 0},
        'jpeg_keep':      {'count': 0, 'before': 0, 'after': 0},
        'alpha_flatten':  {'count': 0, 'before': 0, 'after': 0},
        'downsampled':    {'count': 0, 'before': 0, 'after': 0, 'max_side': max_side},
        'errors':         [],
    }

    total_sheets = len(wb.sheetnames)
    for si, sname in enumerate(wb.sheetnames, 1):
        ws = wb[sname]
        img_count = len(ws._images)
        if img_count == 0:
            continue

        print(f"  [{si}/{total_sheets}] {sname} ({img_count} images)...", end=" ", flush=True)
        new_imgs = []

        for img in ws._images:
            try:
                raw = img._data()
                is_jpeg = raw[:2] == b'\xff\xd8'
                is_png = raw[:4] == b'\x89PNG'
                kb = len(raw) / 1024
                orig_w = getattr(img, 'width', 0) or 0
                orig_h = getattr(img, 'height', 0) or 0

                # ── Open image ──
                pil = PILImage.open(io.BytesIO(raw))

                # ── Downsample ──
                did_downsample = False
                if max_side and pil.size:
                    w, h = pil.size
                    longest = max(w, h)
                    if longest > max_side:
                        scale = max_side / longest
                        nw, nh = int(w * scale), int(h * scale)
                        if pil.mode in ('RGBA', 'LA', 'P'):
                            pil = pil.convert('RGBA')
                            bg = PILImage.new('RGBA', (nw, nh), (255, 255, 255, 255))
                            resized = pil.resize((nw, nh), PILImage.LANCZOS)
                            bg.paste(resized, (0, 0), resized)
                            pil = bg
                        else:
                            pil = pil.resize((nw, nh), PILImage.LANCZOS)
                        did_downsample = True

                # ── Format conversion ──
                if is_png:
                    has_alpha = (pil.mode in ('RGBA', 'LA')
                                 or (pil.mode == 'P' and 'transparency' in pil.info))
                    if has_alpha:
                        if pil.mode != 'RGBA':
                            rgba = pil.convert('RGBA')
                        else:
                            rgba = pil
                        bg = PILImage.new('RGB', rgba.size, (255, 255, 255))
                        bg.paste(rgba, mask=rgba.split()[3])
                        pil = bg
                        bucket = 'alpha_flatten'
                    else:
                        if pil.mode != 'RGB':
                            pil = pil.convert('RGB')
                        bucket = 'png_to_jpeg'

                    out = io.BytesIO()
                    pil.save(out, format='JPEG', quality=quality_default, optimize=True)
                    new_data = out.getvalue()
                    stats[bucket]['count'] += 1
                    stats[bucket]['before'] += len(raw)
                    stats[bucket]['after'] += len(new_data)

                elif is_jpeg and kb > threshold_kb:
                    if pil.mode != 'RGB':
                        pil = pil.convert('RGB')
                    out = io.BytesIO()
                    pil.save(out, format='JPEG', quality=quality_large, optimize=True)
                    new_data = out.getvalue()

                    if len(new_data) < len(raw):
                        stats['jpeg_reencode']['count'] += 1
                        stats['jpeg_reencode']['before'] += len(raw)
                        stats['jpeg_reencode']['after'] += len(new_data)
                    else:
                        new_data = raw
                        stats['jpeg_reencode']['skipped'] += 1
                        stats['jpeg_reencode']['before'] += len(raw)
                        stats['jpeg_reencode']['after'] += len(raw)
                else:
                    # small JPEG that got downsampled → still need to re-encode
                    if did_downsample:
                        if pil.mode != 'RGB':
                            pil = pil.convert('RGB')
                        out = io.BytesIO()
                        pil.save(out, format='JPEG', quality=quality_default, optimize=True)
                        new_data = out.getvalue()
                        stats['jpeg_reencode']['count'] += 1
                        stats['jpeg_reencode']['before'] += len(raw)
                        stats['jpeg_reencode']['after'] += len(new_data)
                    else:
                        new_data = raw
                        stats['jpeg_keep']['count'] += 1
                        stats['jpeg_keep']['before'] += len(raw)
                        stats['jpeg_keep']['after'] += len(raw)

                if did_downsample:
                    stats['downsampled']['count'] += 1

                # ── Rebuild Image object ──
                ns = io.BytesIO(new_data)
                ni = XLImage(ns)
                if did_downsample and orig_w and orig_h and pil.size:
                    nw, nh = pil.size
                    ni.width = int(orig_w * nw / max(orig_w, 1))
                    ni.height = int(orig_h * nh / max(orig_h, 1))
                else:
                    if hasattr(img, 'width'):
                        ni.width = img.width
                    if hasattr(img, 'height'):
                        ni.height = img.height
                if hasattr(img, 'anchor'):
                    ni.anchor = img.anchor
                ni._saved_bytes = new_data
                new_imgs.append(ni)

            except Exception as e:
                stats['errors'].append(f"{sname}#{len(new_imgs)}: {e}")
                new_imgs.append(img)

        ws._images = new_imgs
        print("OK")

    # ── Summary ──
    total_before = sum(s['before'] for s in stats.values() if isinstance(s, dict))
    total_after  = sum(s['after']  for s in stats.values() if isinstance(s, dict))

    print(f"\n{'='*60}")
    print(f"  Results")
    if max_side:
        print(f"  (downsample: longest edge ≤ {max_side}px)")
    print(f"{'='*60}")

    for label, key in [
        (f'Large JPEG→q={quality_large}', 'jpeg_reencode'),
        ('Alpha PNG→JPEG', 'alpha_flatten'),
        ('PNG→JPEG', 'png_to_jpeg'),
        ('Small JPEG kept', 'jpeg_keep'),
    ]:
        s = stats[key]
        if s['count'] == 0:
            continue
        b_mb = s['before'] / (1024 * 1024)
        a_mb = s['after']  / (1024 * 1024)
        ratio = (1 - s['after'] / s['before']) * 100 if s['before'] else 0
        extra = f"(skipped {s['skipped']})" if s.get('skipped') else ""
        print(f"  {label:<22} {s['count']:>4} img  {b_mb:>7.1f}MB → {a_mb:>7.1f}MB  "
              f"{ratio:>5.0f}%↓ {extra}")

    if stats['downsampled']['count']:
        print(f"  Downsampled{'':>14} {stats['downsampled']['count']:>4} img  "
              f"(longest edge > {stats['downsampled']['max_side']}px)")

    print(f"  {'─'*52}")
    print(f"  Total images{'':>12}       {total_before/(1024*1024):>7.1f}MB → "
          f"{total_after/(1024*1024):>7.1f}MB  "
          f"{(1-total_after/total_before)*100 if total_before else 0:>5.0f}%↓")

    src_size = os.path.getsize(src_path)
    print(f"\n  Original file: {src_size/(1024*1024):.1f}MB")

    if stats['errors']:
        print(f"  ⚠ Errors: {len(stats['errors'])}")
        for e in stats['errors'][:5]:
            print(f"    - {e}")
        if len(stats['errors']) > 5:
            print(f"    ... {len(stats['errors'])-5} more")

    if not dry_run:
        print(f"  Saving: {dst_path}")
        wb.save(dst_path)
        dst_size = os.path.getsize(dst_path)
        print(f"  Compressed: {dst_size/(1024*1024):.1f}MB")
        print(f"  Reduced:    {(src_size - dst_size)/(1024*1024):.1f}MB "
              f"({(1 - dst_size/src_size)*100:.0f}%)")
    else:
        print(f"  (dry-run, not saved)")

    wb.close()
    return stats


def main():
    # No args → GUI; with args → CLI
    if len(sys.argv) > 1:
        _main_cli()
    else:
        _main_gui()


def _main_cli():
    parser = argparse.ArgumentParser(
        description='Excel Image Compressor — smart JPEG re-encode + downsampling',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s report.xlsx                       JPEG re-compression only
  %(prog)s report.xlsx --max-side 2048       downsample + compress (recommended)
  %(prog)s report.xlsx --max-side 2048 --quality 80
  %(prog)s report.xlsx --dry-run             stats only, no output
        """)
    parser.add_argument('input', help='Input Excel file (.xlsx)')
    parser.add_argument('-o', '--output', default=None,
                        help='Output path (default: input_compressed.xlsx)')
    parser.add_argument('--max-side', type=int, default=None,
                        help='Max pixel length of longest edge (e.g. 2048). '
                             'Recommended for large microscope images')
    parser.add_argument('--quality', type=int, default=JPEG_QUALITY_LARGE,
                        help=f'JPEG quality for large images 1-100 (default: {JPEG_QUALITY_LARGE})')
    parser.add_argument('--threshold', type=int, default=JPEG_LARGE_THRESHOLD_KB,
                        help=f'File size threshold in KB to trigger re-encode '
                             f'(default: {JPEG_LARGE_THRESHOLD_KB})')
    parser.add_argument('--dry-run', action='store_true',
                        help='Show stats only, do not save')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"Error: file not found - {args.input}")
        sys.exit(1)

    if not args.input.lower().endswith('.xlsx'):
        print(f"Error: only .xlsx files are supported")
        sys.exit(1)

    if args.output is None:
        base, ext = os.path.splitext(args.input)
        args.output = f"{base}_compressed{ext}"

    if args.quality < 1 or args.quality > 100:
        print(f"Error: quality must be between 1 and 100")
        sys.exit(1)

    if args.max_side is not None and args.max_side < 100:
        print(f"Error: --max-side cannot be less than 100")
        sys.exit(1)

    compress_workbook(
        args.input, args.output,
        quality_large=args.quality,
        threshold_kb=args.threshold,
        max_side=args.max_side,
        dry_run=args.dry_run,
    )


# ── GUI ──
def _main_gui():
    from PyQt5.QtWidgets import (
        QApplication, QDialog, QVBoxLayout, QHBoxLayout,
        QLabel, QPushButton, QLineEdit, QSpinBox, QCheckBox,
        QFileDialog, QProgressBar, QTextEdit, QMessageBox, QGroupBox,
    )
    from PyQt5.QtCore import QThread, pyqtSignal

    app = QApplication(sys.argv)
    app.setStyleSheet("""
        QDialog { background: #f5f5f5; }
        QGroupBox { font-weight: bold; border: 1px solid #ddd; border-radius: 6px;
                    margin-top: 8px; padding-top: 14px; background: #fff; }
        QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 4px; }
        QPushButton { padding: 6px 16px; border-radius: 4px; }
        QPushButton#btn_compress { background: #007aff; color: #fff; font-weight: bold;
                                   border: none; padding: 8px 24px; }
        QPushButton#btn_compress:disabled { background: #ccc; }
        QLineEdit, QSpinBox { padding: 4px 6px; border: 1px solid #ccc; border-radius: 4px; }
        QTextEdit { border: 1px solid #ddd; border-radius: 4px; background: #fafafa; }
    """)

    dlg = QDialog()
    dlg.setWindowTitle("Excel Image Compressor")
    dlg.setMinimumWidth(640)
    layout = QVBoxLayout(dlg)
    layout.setSpacing(10)

    # ── File selection ──
    file_group = QGroupBox("File")
    fl = QVBoxLayout(file_group)

    row1 = QHBoxLayout()
    row1.addWidget(QLabel("Excel file:"))
    input_edit = QLineEdit()
    input_edit.setPlaceholderText("Select .xlsx report…")
    input_edit.setReadOnly(True)
    row1.addWidget(input_edit, 1)
    btn_browse = QPushButton("Browse…")
    row1.addWidget(btn_browse)
    fl.addLayout(row1)

    row2 = QHBoxLayout()
    row2.addWidget(QLabel("Output:"))
    output_edit = QLineEdit()
    output_edit.setPlaceholderText("auto: input_compressed.xlsx")
    row2.addWidget(output_edit, 1)
    fl.addLayout(row2)

    layout.addWidget(file_group)

    # ── Settings ──
    param_group = QGroupBox("Settings")
    pl = QHBoxLayout(param_group)

    pl.addWidget(QLabel("Max edge:"))
    max_side_spin = QSpinBox()
    max_side_spin.setRange(0, 99999)
    max_side_spin.setValue(2048)
    max_side_spin.setSpecialValueText("Off")
    max_side_spin.setToolTip("Downsample images whose longest edge exceeds this (px). "
                             "0 = off. Recommended: 2048")
    pl.addWidget(max_side_spin)

    pl.addWidget(QLabel("JPEG quality:"))
    quality_spin = QSpinBox()
    quality_spin.setRange(1, 100)
    quality_spin.setValue(JPEG_QUALITY_LARGE)
    quality_spin.setToolTip("Lower = smaller file. 85 is visually lossless")
    pl.addWidget(quality_spin)

    pl.addWidget(QLabel("Threshold(KB):"))
    threshold_spin = QSpinBox()
    threshold_spin.setRange(50, 5000)
    threshold_spin.setValue(JPEG_LARGE_THRESHOLD_KB)
    threshold_spin.setToolTip("Only re-encode JPEGs larger than this")
    pl.addWidget(threshold_spin)

    dry_check = QCheckBox("Dry-run (stats only)")
    pl.addWidget(dry_check)
    pl.addStretch()
    layout.addWidget(param_group)

    # ── Results ──
    result_group = QGroupBox("Results")
    rl = QVBoxLayout(result_group)

    progress_bar = QProgressBar()
    progress_bar.setVisible(False)
    rl.addWidget(progress_bar)

    result_text = QTextEdit()
    result_text.setReadOnly(True)
    result_text.setMaximumHeight(220)
    result_text.setPlaceholderText("Select a file and click Compress to start…")
    rl.addWidget(result_text)
    layout.addWidget(result_group)

    # ── Buttons ──
    btn_row = QHBoxLayout()
    btn_row.addStretch()
    btn_compress = QPushButton("Compress")
    btn_compress.setObjectName("btn_compress")
    btn_compress.setEnabled(False)
    btn_row.addWidget(btn_compress)
    btn_close = QPushButton("Close")
    btn_row.addWidget(btn_close)
    layout.addLayout(btn_row)

    # ── Background thread ──
    class CompressThread(QThread):
        progress = pyqtSignal(int, int)
        done = pyqtSignal(object, bool)
        log = pyqtSignal(str)

        def __init__(self, src, dst, quality, threshold, max_side, dry_run):
            super().__init__()
            self.src = src
            self.dst = dst
            self.quality = quality
            self.threshold = threshold
            self.max_side = max_side
            self.dry_run = dry_run

        def run(self):
            import io as _io
            from PIL import Image as _PILImage
            _PILImage.MAX_IMAGE_PIXELS = None
            import openpyxl as _xl
            from openpyxl.drawing.image import Image as _XLImage
            import warnings as _w
            _w.filterwarnings('ignore')

            self.log.emit(f"Loading: {self.src}")
            wb = _xl.load_workbook(self.src)

            stats = {
                'png_to_jpeg':   {'count': 0, 'before': 0, 'after': 0},
                'jpeg_reencode': {'count': 0, 'before': 0, 'after': 0, 'skipped': 0},
                'jpeg_keep':     {'count': 0, 'before': 0, 'after': 0},
                'alpha_flatten': {'count': 0, 'before': 0, 'after': 0},
                'downsampled':   {'count': 0, 'before': 0, 'after': 0, 'max_side': self.max_side},
            }
            total_sheets = len(wb.sheetnames)
            ms = self.max_side if self.max_side > 0 else None

            for si, sname in enumerate(wb.sheetnames):
                ws = wb[sname]
                img_count = len(ws._images)
                if img_count == 0:
                    self.progress.emit(si + 1, total_sheets)
                    continue

                self.log.emit(f"  [{si+1}/{total_sheets}] {sname} ({img_count} images)...")
                new_imgs = []

                for img in ws._images:
                    try:
                        raw = img._data()
                        is_jpeg = raw[:2] == b'\xff\xd8'
                        is_png = raw[:4] == b'\x89PNG'
                        kb = len(raw) / 1024
                        orig_w = getattr(img, 'width', 0) or 0
                        orig_h = getattr(img, 'height', 0) or 0

                        pil = _PILImage.open(_io.BytesIO(raw))

                        # Downsample
                        did_downsample = False
                        if ms and pil.size:
                            w, h = pil.size
                            longest = max(w, h)
                            if longest > ms:
                                scale = ms / longest
                                nw, nh = int(w * scale), int(h * scale)
                                if pil.mode in ('RGBA', 'LA', 'P'):
                                    pil = pil.convert('RGBA')
                                    bg = _PILImage.new('RGBA', (nw, nh), (255, 255, 255, 255))
                                    resized = pil.resize((nw, nh), _PILImage.LANCZOS)
                                    bg.paste(resized, (0, 0), resized)
                                    pil = bg
                                else:
                                    pil = pil.resize((nw, nh), _PILImage.LANCZOS)
                                did_downsample = True

                        if is_png:
                            has_alpha = (pil.mode in ('RGBA', 'LA')
                                         or (pil.mode == 'P' and 'transparency' in pil.info))
                            if has_alpha:
                                if pil.mode != 'RGBA':
                                    rgba = pil.convert('RGBA')
                                else:
                                    rgba = pil
                                bg = _PILImage.new('RGB', rgba.size, (255, 255, 255))
                                bg.paste(rgba, mask=rgba.split()[3])
                                pil = bg
                                bucket = 'alpha_flatten'
                            else:
                                if pil.mode != 'RGB':
                                    pil = pil.convert('RGB')
                                bucket = 'png_to_jpeg'
                            out = _io.BytesIO()
                            pil.save(out, format='JPEG', quality=JPEG_QUALITY_DEFAULT, optimize=True)
                            new_data = out.getvalue()
                            stats[bucket]['count'] += 1
                            stats[bucket]['before'] += len(raw)
                            stats[bucket]['after'] += len(new_data)

                        elif is_jpeg and kb > self.threshold:
                            if pil.mode != 'RGB':
                                pil = pil.convert('RGB')
                            out = _io.BytesIO()
                            pil.save(out, format='JPEG', quality=self.quality, optimize=True)
                            new_data = out.getvalue()
                            if len(new_data) < len(raw):
                                stats['jpeg_reencode']['count'] += 1
                                stats['jpeg_reencode']['before'] += len(raw)
                                stats['jpeg_reencode']['after'] += len(new_data)
                            else:
                                new_data = raw
                                stats['jpeg_reencode']['skipped'] += 1
                                stats['jpeg_reencode']['before'] += len(raw)
                                stats['jpeg_reencode']['after'] += len(raw)
                        else:
                            if did_downsample:
                                if pil.mode != 'RGB':
                                    pil = pil.convert('RGB')
                                out = _io.BytesIO()
                                pil.save(out, format='JPEG', quality=JPEG_QUALITY_DEFAULT, optimize=True)
                                new_data = out.getvalue()
                                stats['jpeg_reencode']['count'] += 1
                                stats['jpeg_reencode']['before'] += len(raw)
                                stats['jpeg_reencode']['after'] += len(new_data)
                            else:
                                new_data = raw
                                stats['jpeg_keep']['count'] += 1
                                stats['jpeg_keep']['before'] += len(raw)
                                stats['jpeg_keep']['after'] += len(raw)

                        if did_downsample:
                            stats['downsampled']['count'] += 1

                        ns = _io.BytesIO(new_data)
                        ni = _XLImage(ns)
                        if did_downsample and orig_w and orig_h and pil.size:
                            nw, nh = pil.size
                            ni.width = int(orig_w * nw / max(orig_w, 1))
                            ni.height = int(orig_h * nh / max(orig_h, 1))
                        else:
                            if hasattr(img, 'width'): ni.width = img.width
                            if hasattr(img, 'height'): ni.height = img.height
                        if hasattr(img, 'anchor'): ni.anchor = img.anchor
                        ni._saved_bytes = new_data
                        new_imgs.append(ni)
                    except Exception:
                        new_imgs.append(img)

                ws._images = new_imgs
                self.progress.emit(si + 1, total_sheets)

            self.log.emit("")
            total_before = sum(s['before'] for s in stats.values())
            total_after  = sum(s['after']  for s in stats.values())
            self.log.emit(f"Images: {total_before/(1024*1024):.1f}MB → {total_after/(1024*1024):.1f}MB  "
                          f"({(1-total_after/total_before)*100 if total_before else 0:.0f}%↓)")
            if stats['downsampled']['count']:
                self.log.emit(f"Downsampled: {stats['downsampled']['count']} images")

            src_size = os.path.getsize(self.src)
            self.log.emit(f"Original: {src_size/(1024*1024):.1f}MB")

            if not self.dry_run:
                self.log.emit(f"Saving: {self.dst}")
                wb.save(self.dst)
                dst_size = os.path.getsize(self.dst)
                self.log.emit(f"Compressed: {dst_size/(1024*1024):.1f}MB  "
                              f"↓{(src_size-dst_size)/(1024*1024):.1f}MB "
                              f"({(1-dst_size/src_size)*100:.0f}%)")

            wb.close()
            self.done.emit(stats, self.dry_run)

    thread = None

    # ── Callbacks ──
    def browse():
        path, _ = QFileDialog.getOpenFileName(dlg, "Select Excel Report", "",
                                              "Excel files (*.xlsx);;All files (*)")
        if path:
            input_edit.setText(path)
            base, _ext = os.path.splitext(path)
            output_edit.setText(f"{base}_compressed{_ext}")
            btn_compress.setEnabled(True)

    def start_compress():
        nonlocal thread
        src = input_edit.text()
        if not src or not os.path.exists(src):
            QMessageBox.warning(dlg, "Error", "Please select a valid Excel file.")
            return

        dst = output_edit.text() or None
        if not dst:
            base, _ext = os.path.splitext(src)
            dst = f"{base}_compressed{_ext}"
            output_edit.setText(dst)

        btn_compress.setEnabled(False)
        btn_browse.setEnabled(False)
        result_text.clear()
        progress_bar.setVisible(True)
        progress_bar.setRange(0, 1)

        thread = CompressThread(
            src, dst,
            quality_spin.value(),
            threshold_spin.value(),
            max_side_spin.value(),
            dry_check.isChecked(),
        )

        def on_progress(cur, total):
            progress_bar.setRange(0, total)
            progress_bar.setValue(cur)

        def on_done(_stats, dry_run):
            btn_compress.setEnabled(True)
            btn_browse.setEnabled(True)
            progress_bar.setVisible(False)
            QMessageBox.information(dlg, "Done",
                "Stats only, file not saved." if dry_run else "Compression complete!")

        def on_log(msg):
            result_text.append(msg)

        thread.progress.connect(on_progress)
        thread.done.connect(on_done)
        thread.log.connect(on_log)
        thread.start()

    btn_browse.clicked.connect(browse)
    btn_compress.clicked.connect(start_compress)
    btn_close.clicked.connect(dlg.reject)

    dlg.exec_()
    app.quit()


if __name__ == '__main__':
    main()
