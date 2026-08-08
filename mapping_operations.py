import io
import os
import re
from copy import copy
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Border, Side
from PIL import Image as PILImage

# 本地受信任的高分辨率图片（显微镜切片图常超 Pillow 默认像素上限）：
# 关闭 DecompressionBomb 限制（默认 178,956,970 像素），避免打开大图报错
PILImage.MAX_IMAGE_PIXELS = None
from constants import (
    COL_WIDTH_PX_PER_CHAR, ROW_HEIGHT_PX_PER_PT,
    IMAGE_JPEG_QUALITY, IMAGE_JPEG_QUALITY_LARGE, IMAGE_JPEG_LARGE_THRESHOLD_KB,
    IMAGE_DOWNSAMPLE_MAX_SIDE,
)
from safe_eval import _safe_eval_transform

class MappingOperations:
    """数据/图片/归档/JMP 映射执行逻辑（供 MainWindow 混入）"""
    @staticmethod
    def _archive_blocks_overlap(block1, block2):
        """两个归档区块是否重叠：行区间与列区间都相交才算重叠"""
        r1_min, c1_min, r1_max, c1_max = block1
        r2_min, c2_min, r2_max, c2_max = block2
        return (r1_min <= r2_max and r2_min <= r1_max
                and c1_min <= c2_max and c2_min <= c1_max)
    @staticmethod
    def _block_range_str(block_range):
        min_row, min_col, max_row, max_col = block_range
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    @staticmethod
    def _mapping_key(m):
        """映射去重键：同类型、同目标、同位置视为同一条映射"""
        t = m.get('type')
        if t == 'archive_shift_right':
            return ('archive', m.get('target_sheet'), tuple(m.get('block_range')))
        if t == 'data':
            return ('data', m.get('target_sheet'), tuple(m.get('target_range')),
                    m.get('source_sheet'), m.get('source_range'), m.get('transform'),
                    bool(m.get('clear_target')))
        if t == 'image':
            return ('image', m.get('target_sheet'), m.get('anchor_cell'))
        if t == 'jmp':
            return ('jmp', m.get('target_sheet'), m.get('anchor_cell'))
        if t == 'ocr':
            return ('ocr', m.get('target_sheet'), m.get('mode'),
                    m.get('image_folder'), tuple(m.get('image_list', [])),
                    repr(m.get('target_cells')), tuple(m.get('labels', [])))
        return (t, id(m))
    def _dedupe_mappings(self, mappings):
        """按映射去重键过滤重复项，保留首次出现"""
        seen = set()
        result = []
        for m in mappings:
            key = self._mapping_key(m)
            if key in seen:
                continue
            seen.add(key)
            result.append(m)
        return result
    @staticmethod
    def _extend_block_top_merges(ws, min_row, min_col, max_row, max_col):
        """把覆盖区块首行（或恰好位于首行上方一行，如组表头）的合并单元格向右扩一格。
        与已有合并冲突时跳过，避免合并范围互相覆盖。"""
        targets = []
        for mr in ws.merged_cells.ranges:
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(mr))
            # 行：覆盖首行 或 位于首行上方一行；列：与区块列区间相交
            if not (m_min_row <= min_row <= m_max_row or m_min_row == min_row - 1):
                continue
            if not (m_min_col <= max_col and m_max_col >= min_col):
                continue
            targets.append((str(mr), m_min_row, m_min_col, m_max_row, m_max_col))
        for mr_str, m_min_row, m_min_col, m_max_row, m_max_col in targets:
            # 目标列 m_max_col+1 若已被其他合并占用则跳过
            conflict = False
            for other in ws.merged_cells.ranges:
                if str(other) == mr_str:
                    continue
                o_min_col, o_min_row, o_max_col, o_max_row = range_boundaries(str(other))
                if (m_min_row <= o_max_row and o_min_row <= m_max_row
                        and o_min_col <= m_max_col + 1 <= o_max_col):
                    conflict = True
                    break
            if conflict:
                continue
            ws.unmerge_cells(mr_str)
            ws.merge_cells(start_row=m_min_row, start_column=m_min_col,
                           end_row=m_max_row, end_column=m_max_col + 1)
    @staticmethod
    def _apply_block_borders(ws, block_range):
        """归档后：先对数据块全部单元格加细框线，再对外边界加粗框线"""
        min_row, min_col, max_row, max_col = block_range
        thin = Side(style='thin', color='000000')
        thick = Side(style='thick', color='000000')
        # 1) 全部框线（细线）
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).border = Border(
                    left=thin, right=thin, top=thin, bottom=thin)
        # 2) 粗外框（覆盖外边界）
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                b = cell.border
                cell.border = Border(
                    left=thick if c == min_col else b.left,
                    right=thick if c == max_col else b.right,
                    top=thick if r == min_row else b.top,
                    bottom=thick if r == max_row else b.bottom,
                )
    def execute_mapping(self, ws, mapping, data_ws, data_src_ws=None, jmp_src_ws=None):
        if mapping['type'] == 'data':
            self.apply_data_mapping(ws, mapping, data_src_ws)
        elif mapping['type'] == 'image':
            self.apply_image_mapping(ws, mapping)
        elif mapping['type'] == 'archive_shift_right':
            self.apply_archive_shift_right(ws, mapping, data_ws, data_src_ws)
        elif mapping['type'] == 'jmp':
            self.apply_jmp_mapping(ws, mapping, jmp_src_ws)
        elif mapping['type'] == 'ocr':
            self.apply_ocr_mapping(ws, mapping)
    def _resolve_data_src_ws(self, mapping, data_source_wb):
        """按映射的 source_sheet 从数据源文件解析工作表（优先 data_only 版本）"""
        src_sheet = mapping.get('source_sheet')
        if data_source_wb and src_sheet and src_sheet in data_source_wb.sheetnames:
            return data_source_wb[src_sheet]
        return None
    @staticmethod
    def _excel_error(value):
        """识别 Excel 错误值字符串；非错误值返回 None"""
        if isinstance(value, str) and value in (
            '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A', '#NULL!', '#NUM!'
        ):
            return value
        return None
    def _warn_error_value(self, value, src_addr, dst_addr):
        """若填充值来自数据源且是 Excel 错误值，记录警告供输出时提示"""
        err = self._excel_error(value)
        if err:
            self._fill_warnings.append(f"{dst_addr} <- {src_addr}：数据源错误值 {err}")
        # ---------- 数据转换 ----------
    def _apply_transform(self, value, mapping):
        trans = mapping.get('transform', 'none')
        if trans == 'none' or value is None:
            return value
        try:
            if trans == 'div1000':
                return float(value) / 1000.0
            elif trans == 'strip_letters':
                if isinstance(value, str):
                    return re.sub(r'[a-zA-Z]+$', '', value)
                return value
            elif trans == 'custom':
                expr = mapping.get('transform_expr', '')
                if expr:
                    ok, result = _safe_eval_transform(expr, value)
                    if ok:
                        return result
                    msg = f"自定义表达式计算失败，已保留原值：{expr!r}"
                    if msg not in self._fill_warnings:
                        self._fill_warnings.append(msg)
        except:
            pass
        return value
    def apply_data_mapping(self, ws, mapping, data_src_ws=None):
        t_min_row, t_min_col, t_max_row, t_max_col = mapping['target_range']
        src_ws = data_src_ws if data_src_ws else self.source_wb[mapping['source_sheet']]
        s_min_col, s_min_row, s_max_col, s_max_row = range_boundaries(mapping['source_range'])
        for i in range(t_max_row - t_min_row + 1):
            for j in range(t_max_col - t_min_col + 1):
                src_row = s_min_row + i
                src_col = s_min_col + j
                if src_row <= s_max_row and src_col <= s_max_col:
                    val = src_ws.cell(row=src_row, column=src_col).value
                    val = self._apply_transform(val, mapping)
                    dst_row = t_min_row + i
                    dst_col = t_min_col + j
                    self._warn_error_value(val,
                        f"{src_ws.title}!{get_column_letter(src_col)}{src_row}",
                        f"{ws.title}!{get_column_letter(dst_col)}{dst_row}")
                    ws.cell(row=dst_row, column=dst_col).value = val
    @staticmethod
    def _image_anchor_row_col(img):
        """图片锚点 → (行, 列)。兼容 openpyxl 对象锚点与字符串锚点
        （如 'A6'；新添加的图片在保存前 anchor 保持为字符串）。"""
        anchor = getattr(img, 'anchor', None)
        pos = getattr(anchor, '_from', None)
        if pos is not None:
            return pos.row + 1, pos.col + 1
        if isinstance(anchor, str):
            m = re.match(r'^([A-Z]+)(\d+)$', anchor)
            if m:
                return int(m.group(2)), column_index_from_string(m.group(1))
        return None
    def remove_images_at_anchor(self, ws, anchor):
        m = re.match(r'^([A-Z]+)(\d+)$', anchor)
        target = (int(m.group(2)), column_index_from_string(m.group(1))) if m else None
        if target is None:
            return
        to_remove = []
        for img in ws._images:
            pos = self._image_anchor_row_col(img)
            if pos == target:
                to_remove.append(img)
        for img in to_remove:
            ws._images.remove(img)
    def remove_images_in_region(self, ws, min_row, min_col, max_row, max_col):
        """移除锚点落在指定区域内的全部图片（PBO→MBO 整区清空用）"""
        to_remove = []
        for img in ws._images:
            pos = self._image_anchor_row_col(img)
            if pos is None:
                continue
            r, c = pos
            if min_row <= r <= max_row and min_col <= c <= max_col:
                to_remove.append(img)
        for img in to_remove:
            ws._images.remove(img)
    def clear_region(self, ws, region):
        """清空区域内所有单元格数值与图片（保留样式/合并单元格结构）。
        合并单元格只清左上角（其余为只读 MergedCell），保证区域数据不再显示旧值。"""
        min_row, min_col, max_row, max_col = region
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                try:
                    cell.value = None
                except AttributeError:
                    # MergedCell 值只读，左上角已在上面的循环里清空
                    pass
        self.remove_images_in_region(ws, min_row, min_col, max_row, max_col)
    def _process_image_data(self, img_bytes, rotation, target_w, target_h):
        # 打开前再确认一次：关闭 DecompressionBomb 像素上限，支持超1.79亿像素大图
        PILImage.MAX_IMAGE_PIXELS = None
        # 检测原始格式：JPEG 重编码可激进压缩（q=85 肉眼无差异），
        # PNG 首次转 JPEG 保持高画质（q=90）。
        is_source_jpeg = img_bytes[:2] == b'\xff\xd8'
        source_kb = len(img_bytes) / 1024
        pil_img = PILImage.open(io.BytesIO(img_bytes))
        if rotation != 0:
            pil_img = pil_img.rotate(-rotation, expand=True)
        # 降采样：超大图（最长边 > 2048px）先缩到 2048px 再 resize 到目标尺寸，
        # 避免 2 亿像素原图全解码撑爆内存（207MP → ~4MP → 目标尺寸）
        w, h = pil_img.size
        longest = max(w, h)
        if longest > IMAGE_DOWNSAMPLE_MAX_SIDE:
            scale = IMAGE_DOWNSAMPLE_MAX_SIDE / longest
            nw, nh = int(w * scale), int(h * scale)
            pil_img = pil_img.resize((nw, nh), PILImage.LANCZOS)
        pil_img = pil_img.resize((target_w, target_h), PILImage.LANCZOS)
        # 输出统一为 JPEG：体积小、编码快，降低输出文件与内存占用。
        # JPEG 不支持透明通道，RGBA/调色板透明图先贴到白底再保存。
        has_alpha = (pil_img.mode in ('RGBA', 'LA')
                     or (pil_img.mode == 'P' and 'transparency' in pil_img.info))
        if has_alpha:
            rgba = pil_img.convert('RGBA')
            bg = PILImage.new('RGB', rgba.size, (255, 255, 255))
            bg.paste(rgba, mask=rgba.split()[3])
            pil_img = bg
        else:
            pil_img = pil_img.convert('RGB')
        # 智能选择压缩质量：大 JPEG 原图（>200KB）用 q=85 节省体积，
        # 小 JPEG 和 PNG 转码用 q=90 保证画质
        quality = IMAGE_JPEG_QUALITY
        if is_source_jpeg and source_kb > IMAGE_JPEG_LARGE_THRESHOLD_KB:
            quality = IMAGE_JPEG_QUALITY_LARGE
        out_stream = io.BytesIO()
        pil_img.save(out_stream, format='JPEG',
                     quality=quality, optimize=True)
        out_stream.seek(0)
        return out_stream
    @staticmethod
    def _build_aligned_anchor(anchor_str, alignment, col_width_chars, w_scale, image_width_px):
        """根据对齐方式构建带水平偏移的图片锚点。

        left:   锚定单元格左上角（默认）
        center: 水平居中
        right:  锚定单元格右侧
        """
        if alignment == 'left':
            return anchor_str  # 默认行为，无需偏移

        import re as _re
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker

        m = _re.match(r'^([A-Z]+)(\d+)$', anchor_str)
        if not m:
            return anchor_str

        col_letter, row_num = m.group(1), int(m.group(2))
        col_idx = column_index_from_string(col_letter) - 1  # 0-based

        cell_width_px = int(col_width_chars * COL_WIDTH_PX_PER_CHAR * w_scale)
        if alignment == 'center':
            offset_px = max(0, (cell_width_px - image_width_px) / 2)
        else:  # right
            offset_px = max(0, cell_width_px - image_width_px)

        col_off_emu = int(offset_px * 9525)  # px → EMU (1 px ≈ 9525 EMU @ 96 DPI)
        return AnchorMarker(col=col_idx, colOff=col_off_emu,
                           row=row_num - 1, rowOff=0)

    def apply_image_mapping(self, ws, mapping):
        try:
            anchor = mapping['anchor_cell']
            col_width_chars = mapping.get('col_width_chars', 8.0)
            row_height_pts = mapping.get('row_height_pts', 15.0)
            rotation = mapping.get('rotation', 0.0)
            w_scale = mapping.get('width_scale', 1.0)
            h_scale = mapping.get('height_scale', 1.0)
            alignment = mapping.get('alignment', 'left')
            target_width = int(col_width_chars * COL_WIDTH_PX_PER_CHAR * w_scale)
            target_height = int(row_height_pts * ROW_HEIGHT_PX_PER_PT * h_scale)
            if 'image_path' in mapping:
                if not os.path.exists(mapping['image_path']):
                    raise FileNotFoundError(f"图片文件不存在: {mapping['image_path']}")
                with open(mapping['image_path'], 'rb') as f:
                    img_bytes = f.read()
                processed_stream = self._process_image_data(img_bytes, rotation, target_width, target_height)
                self._image_streams.append(processed_stream)
                self.remove_images_at_anchor(ws, anchor)
                new_img = OpenpyxlImage(processed_stream)
                new_img._saved_bytes = processed_stream.getvalue()
                new_img.anchor = self._build_aligned_anchor(anchor, alignment,
                    col_width_chars, w_scale, target_width)
                ws.add_image(new_img)
                return
            # 配置只记录映射路径：优先按"源Sheet + 锚点位置"从当前数据源
            # 重新读取图片内容，旧配置兼容 image_bytes / image_ref
            img_bytes = None
            if mapping.get('image_src_sheet') and mapping.get('image_src_pos'):
                img_bytes = self._get_image_bytes_by_pos(
                    mapping['image_src_sheet'], mapping['image_src_pos'])
            elif mapping.get('image_bytes'):
                img_bytes = mapping['image_bytes']
            elif 'image_ref' in mapping and self.source_wb:
                img_bytes = self._get_internal_image_bytes(mapping['image_ref'])
            if img_bytes is None:
                raise RuntimeError("映射中缺少图片数据（请先打开 IPQC 数据源或重新选择图片）")
            processed_stream = self._process_image_data(img_bytes, rotation, target_width, target_height)
            self._image_streams.append(processed_stream)
            self.remove_images_at_anchor(ws, anchor)
            new_img = OpenpyxlImage(processed_stream)
            new_img._saved_bytes = processed_stream.getvalue()
            new_img.anchor = self._build_aligned_anchor(anchor, alignment,
                col_width_chars, w_scale, target_width)
            ws.add_image(new_img)
        except Exception as e:
            self._fill_warnings.append(f"图片写入失败 {mapping.get('anchor_cell', '?')}：{e}")
    def _get_internal_image_bytes(self, image_ref):
        """按 (sheet, 序号) 从缓存中取图片字节（避免读取已被 openpyxl 关闭的 ref）"""
        sheet_name, idx = image_ref
        for cached in self.cached_images:
            if cached[0] == sheet_name and cached[1] == idx:
                return cached[3]
        raise RuntimeError(f"未找到图片 {sheet_name}[{idx}]")

    def _get_image_bytes_by_pos(self, sheet_name, pos):
        """按 (sheet, 锚点位置) 从缓存中取图片字节。
        配置只记录来源路径，内容随当前数据源更新，避免旧映射图片缺失。"""
        for cached in self.cached_images:
            if cached[0] == sheet_name and cached[2] == pos:
                return cached[3]
        raise RuntimeError(f"数据源 {sheet_name}!{pos} 未找到图片")

    @staticmethod
    def _jmp_format_ref_row(ws, anchor_row, anchor_col, block_height):
        """JMP 格式参考行：从锚点向上跳过最近的一个数据块（block_height 行），
        返回更早一块的最后一行。找不到更早块时回退到锚点上方最近的非空行。"""
        ref = anchor_row - block_height - 1
        if ref >= 1 and ws.cell(row=ref, column=anchor_col).value is not None:
            return ref
        r = anchor_row - 1
        while r >= 1 and ws.cell(row=r, column=anchor_col).value is None:
            r -= 1
        return r if r >= 1 else anchor_row

    def apply_archive_shift_right(self, ws, mapping, data_ws, data_src_ws=None):
        min_row, min_col, max_row, max_col = mapping['block_range']
        header_rows = mapping.get('header_rows', 1)
        new_headers = mapping.get('new_headers', [])
        merged_to_shift = []
        for mr in ws.merged_cells.ranges:
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(mr))
            if (m_min_row >= min_row and m_max_row <= max_row and
                m_min_col >= min_col and m_max_col <= max_col):
                merged_to_shift.append(str(mr))
        for mr in merged_to_shift:
            ws.unmerge_cells(mr)
        # 右移（纯数值）
        for col in range(max_col, min_col - 1, -1):
            for row in range(min_row, max_row + 1):
                src_cell = data_ws.cell(row=row, column=col)
                dst_cell = ws.cell(row=row, column=col + 1)
                self._warn_error_value(src_cell.value,
                    f"{data_ws.title}!{get_column_letter(col)}{row}",
                    f"{ws.title}!{get_column_letter(col + 1)}{row}")
                dst_cell.value = src_cell.value
                self.copy_cell_style(ws.cell(row=row, column=col), dst_cell)
        for row in range(min_row, max_row + 1):
            ws.cell(row=row, column=min_col).value = None
        for mr in merged_to_shift:
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(mr)
            ws.merge_cells(f"{get_column_letter(m_min_col+1)}{m_min_row}:{get_column_letter(m_max_col+1)}{m_max_row}")
        # 扩展覆盖区块首行（含首行上方一行的组表头）的合并单元格，向右扩一格
        self._extend_block_top_merges(ws, min_row, min_col, max_row, max_col)
        # 填充新表头
        for i in range(header_rows):
            row_idx = min_row + i
            cell = ws.cell(row=row_idx, column=min_col)
            if i < len(new_headers):
                cell.value = new_headers[i]
            self.copy_cell_style(ws.cell(row=row_idx, column=min_col + 1), cell)
        # 填充新数据：默认取同一Sheet"来源列"的数值版（data_ws）写入新列
        data_start_row = min_row + header_rows
        src_col = mapping.get('source_col')
        if src_col:
            try:
                src_col_idx = column_index_from_string(str(src_col).upper())
            except ValueError:
                src_col_idx = None
            if src_col_idx:
                for target_row in range(data_start_row, max_row + 1):
                    src_cell = data_ws.cell(row=target_row, column=src_col_idx)
                    dst_cell = ws.cell(row=target_row, column=min_col)
                    self._warn_error_value(src_cell.value,
                        f"{data_ws.title}!{get_column_letter(src_col_idx)}{target_row}",
                        f"{ws.title}!{get_column_letter(min_col)}{target_row}")
                    dst_cell.value = src_cell.value
                    self.copy_cell_style(ws.cell(row=target_row, column=min_col + 1),
                                         ws.cell(row=target_row, column=min_col))
                # 来源列整体为空：通常是文件未由 Excel 保存过，公式缓存值缺失
                data_rows = max_row - data_start_row + 1
                empty_count = sum(
                    1 for r in range(data_start_row, max_row + 1)
                    if data_ws.cell(row=r, column=src_col_idx).value is None
                )
                if empty_count == data_rows:
                    self._fill_warnings.append(
                        f"{ws.title}!{get_column_letter(min_col)}{min_row}:"
                        f"{get_column_letter(min_col)}{max_row} 的来源列 {src_col} 全部为空，"
                        "可能文件未经 Excel 保存（公式缓存值缺失），请先用 Excel 打开保存后再归档。")
                elif empty_count:
                    self._fill_warnings.append(
                        f"{ws.title}!{get_column_letter(min_col)}{min_row}:"
                        f"{get_column_letter(min_col)}{max_row} 的来源列 {src_col} 有 {empty_count}/{data_rows} 行为空。")
        else:
            # 旧配置兼容：从映射的 source_sheet/source_range 读取
            src_ws = data_src_ws if data_src_ws else self.source_wb[mapping['source_sheet']]
            s_min_col, s_min_row, s_max_col, s_max_row = range_boundaries(mapping['source_range'])
            for i in range(max_row - data_start_row + 1):
                src_row = s_min_row + i
                if src_row > s_max_row:
                    break
                target_row = data_start_row + i
                src_cell = src_ws.cell(row=src_row, column=s_min_col)
                dst_cell = ws.cell(row=target_row, column=min_col)
                self._warn_error_value(src_cell.value,
                    f"{src_ws.title}!{get_column_letter(s_min_col)}{src_row}",
                    f"{ws.title}!{get_column_letter(min_col)}{target_row}")
                dst_cell.value = src_cell.value
                self.copy_cell_style(ws.cell(row=target_row, column=min_col + 1),
                                     ws.cell(row=target_row, column=min_col))
        # 归档完成后：对数据块（含新列）执行"全部框线 + 粗外框"
        self._apply_block_borders(ws, (min_row, min_col, max_row, max_col + 1))
    def apply_jmp_mapping(self, ws, mapping, src_data_ws=None):
        anchor = mapping['anchor_cell']
        col_letter = ''.join(filter(str.isalpha, anchor))
        row_num = int(''.join(filter(str.isdigit, anchor)))
        start_row = row_num
        start_col = column_index_from_string(col_letter)
        # 表头列数与内容：数据一律从 锚点列 + 表头列数 之后开始张贴
        # header_cols 在配置导入时已标准化为字符串列表（utils.normalize_mappings）
        headers = [str(h) for h in mapping.get('header_cols', [])]
        header_count = len(headers)
        data_start_col = start_col + header_count
        src_sheet_name = mapping['source_sheet']
        src_range = mapping['source_range']
        merge = mapping['merge_columns']
        # 源数据必须取自映射指定的源Sheet的数值版（data_only），不能用目标Sheet代替
        src_ws = src_data_ws if src_data_ws is not None else self.template_wb[src_sheet_name]
        s_min_col, s_min_row, s_max_col, s_max_row = range_boundaries(src_range)
        data_rows = s_max_row - s_min_row + 1
        data_cols = s_max_col - s_min_col + 1
        # 参考行样式：从锚点向上跳过最近的一个数据块（块高=本次写入行数），
        # 取更早一块的格式；找不到时回退到锚点上一行
        block_height = data_rows * data_cols if (merge and data_cols > 1) else data_rows
        ref_row = self._jmp_format_ref_row(ws, start_row, start_col, block_height)
        if merge and data_cols > 1:
            values = []
            for r in range(s_min_row, s_max_row + 1):
                for c in range(s_min_col, s_max_col + 1):
                    val = src_ws.cell(row=r, column=c).value
                    values.append(val)
            row_count = len(values)
            for row_offset in range(row_count):
                current_row = start_row + row_offset
                # 写入表头
                for idx, header_text in enumerate(headers):
                    cell = ws.cell(row=current_row, column=start_col + idx)
                    cell.value = header_text
                # 写入数据
                val = values[row_offset]
                ws.cell(row=current_row, column=data_start_col, value=val)
                # 复制格式（整行）
                for col_idx in range(start_col, data_start_col + 1):
                    ref_cell = ws.cell(row=ref_row, column=col_idx)
                    target_cell = ws.cell(row=current_row, column=col_idx)
                    self.copy_cell_style(ref_cell, target_cell)
        else:
            row_count = data_rows
            target_data_cols = data_cols
            for row_offset in range(row_count):
                current_row = start_row + row_offset
                # 写入表头
                for idx, header_text in enumerate(headers):
                    cell = ws.cell(row=current_row, column=start_col + idx)
                    cell.value = header_text
                # 写入数据
                for c in range(target_data_cols):
                    src_val = src_ws.cell(row=s_min_row + row_offset, column=s_min_col + c).value
                    ws.cell(row=current_row, column=data_start_col + c, value=src_val)
                # 复制格式
                for col_idx in range(start_col, data_start_col + target_data_cols):
                    ref_cell = ws.cell(row=ref_row, column=col_idx)
                    target_cell = ws.cell(row=current_row, column=col_idx)
                    self.copy_cell_style(ref_cell, target_cell)
    def apply_ocr_mapping(self, ws, mapping):
        """执行 OCR 映射：对每张切片图像做 OCR，将提取的量测值写入目标单元格。

        支持 single（单值）和 labeled（多标注值）两种模式。
        OCR 失败的单元格保留为空并记录 warning。
        """
        import os as _os
        from PyQt5.QtWidgets import QApplication
        from ocr_engine import ocr_batch

        image_folder = mapping.get('image_folder', '')
        image_list = mapping.get('image_list', [])
        mode = mapping.get('mode', 'first_number')
        labels = mapping.get('labels')
        roi = mapping.get('roi')
        preprocess = mapping.get('preprocess', 'none')
        lang = mapping.get('lang', 'ch')
        expr = mapping.get('expr', '')
        target_cells = mapping.get('target_cells', [])

        # 构建绝对路径
        image_paths = [_os.path.join(image_folder, p) for p in image_list]

        # 检查文件是否存在
        missing = [p for p in image_paths if not _os.path.exists(p)]
        if missing:
            self._fill_warnings.append(
                f"OCR 映射（{ws.title}）：{len(missing)}/{len(image_paths)} 个文件缺失，"
                f"如 {_os.path.basename(missing[0])}")

        # 执行批量 OCR
        try:
            def _tick(cur, total):
                # 批量 OCR 期间刷新进度弹窗，避免界面看起来卡死
                QApplication.processEvents()
            results = ocr_batch(
                image_paths, roi=roi, preprocess=preprocess, lang=lang,
                mode=mode if mode != 'labeled' else 'labeled',
                labels=labels, expr=expr, progress_callback=_tick)
        except Exception as e:
            self._fill_warnings.append(
                f"OCR 批量识别失败（{ws.title}）：{e}")
            return

        # 写入目标单元格
        for idx, result in enumerate(results):
            img_name = result.get('_image', '?')
            if '_error' in result:
                err = result['_error']
                self._fill_warnings.append(
                    f"OCR 失败 {img_name}：{err}")
                continue

            if idx >= len(target_cells):
                break

            tc = target_cells[idx]

            if mode == 'labeled' and labels:
                # 多值模式：每张图对应一行 target_cells（list of [row,col]）
                if isinstance(tc, list) and len(tc) > 0 and isinstance(tc[0], (list, tuple)):
                    for li, label in enumerate(labels):
                        if li >= len(tc):
                            break
                        row, col = tc[li][0], tc[li][1]
                        val = result.get(label)
                        cell_addr = f"{ws.title}!{get_column_letter(col)}{row}"
                        if val is None:
                            self._fill_warnings.append(
                                f"{cell_addr} <- {img_name}/{label}：未识别到数值")
                        ws.cell(row=row, column=col).value = val
                elif isinstance(tc, (list, tuple)) and len(tc) >= 1:
                    # 单行模式：取第一个标签值写入单个单元格
                    row, col = tc[0], tc[1]
                    first_label = labels[0] if labels else None
                    val = result.get(first_label) if first_label else result.get('value')
                    ws.cell(row=row, column=col).value = val
            else:
                # 单值/全部数值/自定义模式：每张图对应一个 target_cell
                if isinstance(tc, (list, tuple)):
                    row, col = tc[0], tc[1]
                else:
                    row, col = tc, tc  # fallback, shouldn't happen
                cell_addr = f"{ws.title}!{get_column_letter(col)}{row}"
                if mode == 'all_numbers':
                    vals = result.get('values', [])
                    if not vals:
                        self._fill_warnings.append(
                            f"{cell_addr} <- {img_name}：未识别到数值")
                    elif len(vals) == 1:
                        ws.cell(row=row, column=col).value = vals[0]
                    else:
                        # 多个数值：目标只有单格时以逗号拼接写入
                        ws.cell(row=row, column=col).value = \
                            ', '.join(str(v) for v in vals)
                else:
                    val = result.get('value')
                    if val is None:
                        self._fill_warnings.append(
                            f"{cell_addr} <- {img_name}：未识别到数值")
                    ws.cell(row=row, column=col).value = val

    @staticmethod
    def copy_cell_style(src, dst):
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
