import os
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QWidget, QLabel, QPushButton,
    QFormLayout, QLineEdit, QComboBox, QDialogButtonBox, QAbstractItemView,
    QSpinBox, QGroupBox, QRadioButton, QButtonGroup, QCheckBox, QScrollArea,
    QTableWidget, QTableWidgetItem, QListWidget, QListWidgetItem, QDoubleSpinBox,
    QProgressBar, QApplication, QFileDialog,
)
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtCore import Qt, QSize, QTimer
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
from constants import (
    COL_WIDTH_PX_PER_CHAR, ROW_HEIGHT_PX_PER_PT,
    DEFAULT_COL_WIDTH_CHARS, DEFAULT_COL_WIDTH_PX,
    DEFAULT_ROW_HEIGHT_PTS, DEFAULT_ROW_HEIGHT_PX,
    MAX_PREVIEW_ROWS, MAX_PREVIEW_COLS,
)
from version_finder import detect_version, suggest_files

# ================== 图片设置对话框 ==================
class ImageSetupDialog(QDialog):
    def __init__(self, parent=None, default_col_width=None, default_row_height=None):
        super().__init__(parent)
        self.setWindowTitle("图片设置")
        layout = QFormLayout(self)

        self.col_width = QDoubleSpinBox()
        self.col_width.setRange(1.0, 1000.0)
        self.col_width.setDecimals(1)
        self.col_width.setValue(default_col_width if default_col_width else 8.0)
        layout.addRow("列宽(字符):", self.col_width)

        self.row_height = QDoubleSpinBox()
        self.row_height.setRange(1.0, 2000.0)
        self.row_height.setDecimals(1)
        self.row_height.setValue(default_row_height if default_row_height else 15.0)
        layout.addRow("行高(磅):", self.row_height)

        self.rotation = QDoubleSpinBox()
        self.rotation.setRange(-360.0, 360.0)
        self.rotation.setDecimals(1)
        self.rotation.setValue(0.0)
        layout.addRow("旋转角度(正顺时针):", self.rotation)

        self.width_scale = QDoubleSpinBox()
        self.width_scale.setRange(0.1, 1.0)
        self.width_scale.setSingleStep(0.1)
        self.width_scale.setValue(1.0)
        layout.addRow("宽度缩放比例:", self.width_scale)

        self.height_scale = QDoubleSpinBox()
        self.height_scale.setRange(0.1, 1.0)
        self.height_scale.setSingleStep(0.1)
        self.height_scale.setValue(1.0)
        layout.addRow("高度缩放比例:", self.height_scale)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_values(self):
        return (self.col_width.value(), self.row_height.value(),
                self.rotation.value(),
                self.width_scale.value(), self.height_scale.value())


# ================== 版本号查找对话框 ==================
class VersionFinderDialog(QDialog):
    """从用户指定的档案中识别 Process control / ERS / VSR / MCO 版本号"""
    FIELDS = [
        ("Process control rev.", "process_control",
         "匹配 'Table : Process Control Rev x.xx'"),
        ("ERS rev.", "ers", "文件名 RevN 或正文"),
        ("VSR rev.", "vsr", "文件名 RevN 或正文"),
        ("MCO rev.", "mco", "文件名末尾 -NN 或 RevN"),
    ]

    def __init__(self, parent=None, suggestions=None, folder_keyword='BUF'):
        super().__init__(parent)
        self.setWindowTitle("查找版本号")
        self.setMinimumWidth(760)
        suggestions = suggestions or {}
        self.folder_keyword = folder_keyword
        self._detected = {}

        layout = QVBoxLayout(self)
        hint = QLabel(
            "为每个字段选择对应档案，程序会自动识别版本号并写入 Summary；"
            "识别失败的项目将跳过更新，请在输出报告中手动修改。")
        hint.setWordWrap(True)
        layout.addWidget(hint)

        folder_btn = QPushButton("选择档案文件夹自动识别…")
        folder_btn.clicked.connect(self.pick_folder)
        layout.addWidget(folder_btn)

        self.rows = {}
        for label, key, tip in self.FIELDS:
            row = QWidget()
            hl = QHBoxLayout(row)
            hl.setContentsMargins(0, 0, 0, 0)
            lbl = QLabel(label)
            lbl.setMinimumWidth(140)
            edit = QLineEdit()
            edit.setReadOnly(True)
            edit.setPlaceholderText(tip)
            btn = QPushButton("浏览…")
            status = QLabel("未识别")
            status.setMinimumWidth(150)
            btn.clicked.connect(lambda _, k=key: self.browse(k))
            hl.addWidget(lbl)
            hl.addWidget(edit, 1)
            hl.addWidget(btn)
            hl.addWidget(status)
            layout.addWidget(row)
            self.rows[key] = (edit, status)
            suggested = suggestions.get(key)
            if suggested:
                edit.setText(suggested)
                self.detect(key, suggested)

        layout.addStretch()
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def browse(self, key):
        current = self.rows[key][0].text()
        start_dir = os.path.dirname(current) if current else ""
        path, _ = QFileDialog.getOpenFileName(
            self, "选择档案", start_dir,
            "PDF/Excel (*.pdf *.xlsx);;所有文件 (*)")
        if not path:
            return
        self.rows[key][0].setText(path)
        self.detect(key, path)

    def pick_folder(self):
        """选择一个档案文件夹（含 ERS/VSR/MCO 子文件夹），自动填充并识别"""
        current = self.rows['ers'][0].text()
        start_dir = os.path.dirname(current) if current else ""
        folder = QFileDialog.getExistingDirectory(self, "选择档案文件夹", start_dir)
        if not folder:
            return
        sugg = suggest_files(folder, self.folder_keyword)
        for key in ('ers', 'vsr', 'mco'):
            if key in sugg:
                self.rows[key][0].setText(sugg[key])
                self.detect(key, sugg[key])
        ers_path = self.rows['ers'][0].text()
        if ers_path:
            self.rows['process_control'][0].setText(ers_path)
            self.detect('process_control', ers_path)

    def detect(self, key, path):
        value = detect_version(key, path)
        self._detected[key] = value
        _, status = self.rows[key]
        if value is not None:
            status.setText(f"识别到: {value}")
            status.setStyleSheet("color: #1e7e34; font-weight: bold;")
        else:
            status.setText("匹配失败，跳过更新")
            status.setStyleSheet("color: #b00020; font-weight: bold;")

    def get_results(self):
        """返回 {字段: (档案路径, 识别值或 None)}"""
        results = {}
        for key, (edit, _) in self.rows.items():
            path = edit.text().strip()
            if not path:
                continue
            results[key] = (path, self._detected.get(key))
        return results


# ================== 数据转换规则对话框 ==================
class TransformDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("数据转换规则")
        layout = QVBoxLayout(self)

        self.btn_group = QButtonGroup(self)
        self.radio_none = QRadioButton("无转换")
        self.radio_div1000 = QRadioButton("除以1000 (ms→s)")
        self.radio_strip = QRadioButton("去除末尾字母")
        self.radio_custom = QRadioButton("自定义表达式")
        self.btn_group.addButton(self.radio_none, 0)
        self.btn_group.addButton(self.radio_div1000, 1)
        self.btn_group.addButton(self.radio_strip, 2)
        self.btn_group.addButton(self.radio_custom, 3)
        self.radio_none.setChecked(True)

        layout.addWidget(self.radio_none)
        layout.addWidget(self.radio_div1000)
        layout.addWidget(self.radio_strip)
        layout.addWidget(self.radio_custom)

        self.custom_expr = QLineEdit()
        self.custom_expr.setPlaceholderText("输入表达式，如 x/1000（支持四则运算与 abs/round/min/max/int/float/str/len/sum）")
        self.custom_expr.setEnabled(False)
        self.radio_custom.toggled.connect(lambda checked: self.custom_expr.setEnabled(checked))
        layout.addWidget(self.custom_expr)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_transform(self):
        if self.radio_none.isChecked():
            return 'none', ''
        elif self.radio_div1000.isChecked():
            return 'div1000', ''
        elif self.radio_strip.isChecked():
            return 'strip_letters', ''
        else:
            return 'custom', self.custom_expr.text()


# ================== 归档配置对话框（动态输入框） ==================
class ArchiveConfigDialog(QDialog):
    def __init__(self, template_range=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("归档配置")
        self.setMinimumSize(600, 500)
        self.header_inputs = []  # 动态生成的表头输入框

        main_layout = QVBoxLayout(self)
        # 模板区域信息
        if template_range:
            min_row, min_col, max_row, max_col = template_range
            rows = max_row - min_row + 1
            info_text = f"📌 已选首列范围：{rows} 行（{get_column_letter(min_col)}{min_row}:{get_column_letter(min_col)}{max_row}）"
        else:
            info_text = "⚠️ 未获取到模板区域，请重新框选"
        info_label = QLabel(info_text)
        info_label.setStyleSheet("font-weight: bold; color: #2c3e50; background-color: #ecf0f1; padding: 5px;")
        main_layout.addWidget(info_label)

        # 表头行数选择（改变时动态生成输入框）
        form = QFormLayout()
        self.header_rows_spin = QSpinBox()
        self.header_rows_spin.setMinimum(1)
        self.header_rows_spin.setValue(1)
        self.header_rows_spin.valueChanged.connect(self.on_header_rows_changed)
        form.addRow("表头行数:", self.header_rows_spin)
        main_layout.addLayout(form)

        # 新数据来源列（默认 J，即 Current Config 列；取同一Sheet的数值版）
        src_form = QFormLayout()
        self.source_col_edit = QLineEdit("J")
        self.source_col_edit.setPlaceholderText("例如 J")
        self.source_col_edit.setMaxLength(3)
        src_form.addRow("新数据来源列（同一Sheet）:", self.source_col_edit)
        main_layout.addLayout(src_form)

        # 动态输入框容器
        self.header_group = QGroupBox("新表头内容（每行一个输入框）")
        self.header_layout = QVBoxLayout(self.header_group)
        self.header_scroll = QScrollArea()
        self.header_scroll.setWidgetResizable(True)
        self.header_scroll.setWidget(self.header_group)
        main_layout.addWidget(self.header_scroll)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        main_layout.addWidget(buttons)

        self.on_header_rows_changed(1)  # 初始化输入框

    def on_header_rows_changed(self, value):
        # 清空并重新生成输入框
        for i in reversed(range(self.header_layout.count())):
            widget = self.header_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)
        self.header_inputs.clear()
        for i in range(value):
            edit = QLineEdit()
            edit.setPlaceholderText(f"第{i+1}行表头")
            self.header_layout.addWidget(edit)
            self.header_inputs.append(edit)

    def get_selection(self):
        header_rows = self.header_rows_spin.value()
        headers = [edit.text() for edit in self.header_inputs]
        source_col = self.source_col_edit.text().strip().upper()
        return header_rows, headers, source_col


# ================== JMP配置对话框（动态列数） ==================
class JMPConfigDialog(QDialog):
    def __init__(self, template_wb, anchor_cell, parent=None):
        super().__init__(parent)
        self.setWindowTitle("JMP数据区配置")
        self.setMinimumSize(600, 450)
        self.template_wb = template_wb
        self.anchor_cell = anchor_cell
        self.header_inputs = []  # 动态生成的表头输入框

        layout = QVBoxLayout(self)

        info_label = QLabel(f"📍 锚点单元格: {anchor_cell}")
        info_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        layout.addWidget(info_label)

        # 表头列数选择
        header_form = QFormLayout()
        self.header_cols_spin = QSpinBox()
        self.header_cols_spin.setMinimum(1)
        self.header_cols_spin.setValue(2)
        self.header_cols_spin.valueChanged.connect(self.on_header_cols_changed)
        header_form.addRow("表头列数:", self.header_cols_spin)
        layout.addLayout(header_form)

        # 动态输入框容器
        self.header_group = QGroupBox("表头内容（每列一个输入框）")
        self.header_layout = QVBoxLayout(self.header_group)
        self.header_scroll = QScrollArea()
        self.header_scroll.setWidgetResizable(True)
        self.header_scroll.setWidget(self.header_group)
        layout.addWidget(self.header_scroll)

        # 源数据区域
        source_group = QGroupBox("源数据区域（从模板Sheet中选择）")
        source_layout = QFormLayout(source_group)
        self.sheet_combo = QComboBox()
        self.sheet_combo.addItems(template_wb.sheetnames)
        source_layout.addRow("源Sheet:", self.sheet_combo)

        range_layout = QHBoxLayout()
        self.range_edit = QLineEdit()
        self.range_edit.setPlaceholderText("例如 C2:E10，或点击右侧按钮框选")
        range_layout.addWidget(self.range_edit)
        btn_select = QPushButton("框选区域")
        btn_select.clicked.connect(self.select_range)
        range_layout.addWidget(btn_select)
        source_layout.addRow("源区域:", range_layout)
        layout.addWidget(source_group)

        self.merge_cols_check = QCheckBox("将多列数据拼接为单列（先行后列）")
        layout.addWidget(self.merge_cols_check)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.on_header_cols_changed(2)  # 初始化输入框

    def on_header_cols_changed(self, value):
        # 清空并重新生成输入框
        for i in reversed(range(self.header_layout.count())):
            widget = self.header_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)
        self.header_inputs.clear()
        for i in range(value):
            edit = QLineEdit()
            edit.setPlaceholderText(f"第{i+1}列表头")
            self.header_layout.addWidget(edit)
            self.header_inputs.append(edit)

    def select_range(self):
        sheet_name = self.sheet_combo.currentText()
        if not sheet_name:
            return
        dlg = SourceSelectDialog(self.template_wb, self)
        dlg.sheet_combo.setCurrentText(sheet_name)
        if dlg.exec_():
            sel_sheet, sel_range = dlg.get_selection()
            if sel_range:
                self.range_edit.setText(sel_range)

    def get_config(self):
        headers = [edit.text() for edit in self.header_inputs]
        return {
            'header_cols': headers,
            'source_sheet': self.sheet_combo.currentText(),
            'source_range': self.range_edit.text(),
            'merge_columns': self.merge_cols_check.isChecked()
        }


# ================== 源数据选择对话框 ==================
class SourceSelectDialog(QDialog):
    def __init__(self, source_wb, parent=None, default_sheet=None):
        super().__init__(parent)
        self.setWindowTitle("选择数据源区域")
        self.setMinimumSize(800, 600)
        self.source_wb = source_wb
        self.selected_range = None

        main_layout = QVBoxLayout(self)
        top_layout = QHBoxLayout()
        top_layout.addWidget(QLabel("源Sheet:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.addItems(source_wb.sheetnames)
        if default_sheet and default_sheet in source_wb.sheetnames:
            self.sheet_combo.setCurrentText(default_sheet)
        self.sheet_combo.currentIndexChanged.connect(self.on_sheet_changed)
        top_layout.addWidget(self.sheet_combo)
        top_layout.addStretch()
        main_layout.addLayout(top_layout)

        self.table = QTableWidget()
        self.table.setSelectionMode(QAbstractItemView.ContiguousSelection)
        self.table.itemSelectionChanged.connect(self.on_selection_changed)
        main_layout.addWidget(self.table)

        bottom_layout = QHBoxLayout()
        self.range_label = QLabel("未选择区域")
        bottom_layout.addWidget(self.range_label)
        bottom_layout.addStretch()
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        bottom_layout.addWidget(buttons)
        main_layout.addLayout(bottom_layout)

        self.load_sheet(self.sheet_combo.currentText())

    def on_sheet_changed(self, index):
        self.load_sheet(self.sheet_combo.currentText())

    def load_sheet(self, sheet_name):
        ws = self.source_wb[sheet_name]
        self.table.clear()
        self.table.clearSpans()
        real_max_row = ws.max_row
        real_max_col = ws.max_column
        if real_max_row > 200:
            for r in range(real_max_row, 0, -1):
                if any(cell.value is not None for cell in ws[r]):
                    real_max_row = r
                    break
            real_max_row += 2
        if real_max_col > 30:
            for c in range(real_max_col, 0, -1):
                if any(ws.cell(row=r, column=c).value is not None for r in range(1, real_max_row+1)):
                    real_max_col = c
                    break
            real_max_col += 2
        max_rows = min(real_max_row, MAX_PREVIEW_ROWS)
        max_cols = min(real_max_col, MAX_PREVIEW_COLS)

        self.table.setRowCount(max_rows)
        self.table.setColumnCount(max_cols)

        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row > max_rows or min_col > max_cols:
                continue
            span_rows = min(max_row, max_rows) - min_row + 1
            span_cols = min(max_col, max_cols) - min_col + 1
            self.table.setSpan(min_row-1, min_col-1, span_rows, span_cols)

        # 稀疏遍历：只处理真实存在的单元格，避免大表物化空单元格
        for (row_idx, col_idx), cell in ws._cells.items():
            if row_idx > max_rows or col_idx > max_cols:
                continue
            if cell.value is not None:
                item = QTableWidgetItem(str(cell.value))
                self.table.setItem(row_idx - 1, col_idx - 1, item)

        self._apply_uniform_sizes(ws, max_cols, max_rows)
        self.selected_range = None
        self.range_label.setText("未选择区域")

    def _apply_uniform_sizes(self, ws, max_cols, max_rows):
        for col in range(1, max_cols + 1):
            width_chars = self._column_width_chars(ws, col)
            width = int(width_chars * COL_WIDTH_PX_PER_CHAR)
            self.table.setColumnWidth(col - 1, width)
        for row in range(1, max_rows + 1):
            if row in ws.row_dimensions and ws.row_dimensions[row].height:
                height = int(ws.row_dimensions[row].height * ROW_HEIGHT_PX_PER_PT)
            else:
                default_pts = ws.sheet_format.defaultRowHeight or DEFAULT_ROW_HEIGHT_PTS
                height = int(default_pts * ROW_HEIGHT_PX_PER_PT)
            self.table.setRowHeight(row - 1, height)

    @staticmethod
    def _column_width_chars(ws, col_idx):
        """解析某列宽度（字符数），兼容 openpyxl 不展开的区间列定义。
        Excel 常写成 <col min="3" max="13" width="33"/>，openpyxl 只保留
        首列索引，其余列查不到，需要遍历全部维度按 min/max 区间匹配。
        未定义列退回 sheet 默认列宽。"""
        letter = get_column_letter(col_idx)
        if letter in ws.column_dimensions:
            dim = ws.column_dimensions[letter]
            if dim.width:
                return dim.width
        for dim in ws.column_dimensions.values():
            lo = getattr(dim, 'min', None) or 0
            hi = getattr(dim, 'max', None) or lo
            if lo <= col_idx <= hi and dim.width:
                return dim.width
        return ws.sheet_format.defaultColWidth or DEFAULT_COL_WIDTH_CHARS

    def on_selection_changed(self):
        indexes = self.table.selectedIndexes()
        if not indexes:
            self.selected_range = None
            self.range_label.setText("未选择区域")
            return
        top = min(idx.row() for idx in indexes) + 1
        left = min(idx.column() for idx in indexes) + 1
        bottom = max(idx.row() for idx in indexes) + 1
        right = max(idx.column() for idx in indexes) + 1
        self.selected_range = (top, left, bottom, right)
        addr = f"{get_column_letter(left)}{top}:{get_column_letter(right)}{bottom}"
        self.range_label.setText(addr)

    def get_selection(self):
        sheet_name = self.sheet_combo.currentText()
        if self.selected_range:
            min_row, min_col, max_row, max_col = self.selected_range
            range_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        else:
            range_str = ""
        return sheet_name, range_str


# ================== 内部图片选择对话框 ==================
class InternalImageSelectDialog(QDialog):
    def __init__(self, cached_images, parent=None, default_sheet=None):
        super().__init__(parent)
        self.setWindowTitle("选择图片（可多选）")
        self.setMinimumSize(960, 600)
        self.cached_images = cached_images

        main_layout = QVBoxLayout(self)

        # 顶部：搜索 + Sheet 筛选
        filter_layout = QHBoxLayout()
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("🔍 搜索位置（如 C6、H7）或 Sheet 名")
        self.search_edit.textChanged.connect(self.refresh_list)
        filter_layout.addWidget(self.search_edit, 1)
        filter_layout.addWidget(QLabel("Sheet:"))
        self.sheet_filter = QComboBox()
        sheets = sorted(list(set(img[0] for img in cached_images)))
        self.sheet_filter.addItem("全部")
        self.sheet_filter.addItems(sheets)
        if default_sheet and default_sheet in sheets:
            self.sheet_filter.setCurrentText(default_sheet)
        self.sheet_filter.currentIndexChanged.connect(self.refresh_list)
        filter_layout.addWidget(self.sheet_filter)
        main_layout.addLayout(filter_layout)

        # 中间：左列表（缩略图）+ 右预览
        body = QHBoxLayout()
        self.list = QListWidget()
        self.list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.list.setIconSize(QSize(64, 64))
        self.list.setSpacing(2)
        self.list.itemSelectionChanged.connect(self.update_count)
        self.list.currentItemChanged.connect(self.update_preview)
        self.list.setWordWrap(True)
        body.addWidget(self.list, 3)

        self.preview = QLabel("在左侧选择图片查看预览")
        self.preview.setAlignment(Qt.AlignCenter)
        self.preview.setMinimumSize(320, 320)
        self.preview.setStyleSheet("border:1px solid #cccccc; background:#fafafa;")
        body.addWidget(self.preview, 2)
        main_layout.addLayout(body, 1)

        # 缩略图生成进度条（图片多时可见）
        self.thumb_progress = QProgressBar()
        self.thumb_progress.setVisible(False)
        self.thumb_progress.setFormat("正在生成缩略图 %v/%m，请稍等…")
        main_layout.addWidget(self.thumb_progress)

        # 底部：选择信息 + 按钮
        bottom = QHBoxLayout()
        self.count_label = QLabel("已选 0 张")
        bottom.addWidget(self.count_label)
        self.btn_select_all = QPushButton("全选当前列表")
        self.btn_select_all.clicked.connect(self.select_all)
        bottom.addWidget(self.btn_select_all)
        self.btn_clear = QPushButton("清空选择")
        self.btn_clear.clicked.connect(self.clear_selection)
        bottom.addWidget(self.btn_clear)
        bottom.addStretch()
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        bottom.addWidget(btn_box)
        main_layout.addLayout(bottom)

        # 延迟到对话框显示后再生成缩略图，进度条才可见
        QTimer.singleShot(0, self.refresh_list)

    def _thumbnail(self, img_data):
        pix = QPixmap()
        if img_data and pix.loadFromData(img_data):
            return QIcon(pix.scaled(QSize(64, 64), Qt.KeepAspectRatio, Qt.SmoothTransformation))
        return QIcon()

    def refresh_list(self):
        """按搜索词和 Sheet 筛选重建列表"""
        search = self.search_edit.text().strip().lower()
        sheet = self.sheet_filter.currentText()
        visible = [ci for ci, img in enumerate(self.cached_images)
                   if (sheet == "全部" or img[0] == sheet)
                   and (not search or search in img[2].lower() or search in img[0].lower())]
        self.list.clear()
        total = len(visible)
        show_progress = total > 30
        if show_progress:
            self.thumb_progress.setRange(0, total)
            self.thumb_progress.setVisible(True)
            QApplication.processEvents()
        for n, ci in enumerate(visible, 1):
            sh, idx, pos, data, w, h = self.cached_images[ci]
            item = QListWidgetItem(self._thumbnail(data), f"{pos}  ({w}×{h})")
            item.setData(Qt.UserRole, ci)  # 缓存图片列表中的原始索引
            item.setToolTip(f"{sh}  {pos}  {w}×{h}")
            self.list.addItem(item)
            if show_progress and (n % 20 == 0 or n == total):
                self.thumb_progress.setValue(n)
                QApplication.processEvents()
        if show_progress:
            self.thumb_progress.setVisible(False)
        self.update_count()

    def select_all(self):
        for i in range(self.list.count()):
            self.list.item(i).setSelected(True)
        self.update_count()

    def clear_selection(self):
        self.list.clearSelection()
        self.update_count()

    def update_count(self):
        n = len(self.list.selectedItems())
        self.count_label.setText(f"已选 {n} 张")

    def update_preview(self, current, previous=None):
        if current is None:
            self.preview.setText("在左侧选择图片查看预览")
            self.preview.setPixmap(QPixmap())
            return
        ci = current.data(Qt.UserRole)
        if ci is None or not (0 <= ci < len(self.cached_images)):
            return
        data = self.cached_images[ci][3]
        pix = QPixmap()
        if data and pix.loadFromData(data):
            self.preview.setPixmap(
                pix.scaled(QSize(360, 360), Qt.KeepAspectRatio, Qt.SmoothTransformation))
            self.preview.setText("")
        else:
            self.preview.setText("无法预览该图片")

    def get_selected_images(self):
        result = []
        for item in self.list.selectedItems():
            ci = item.data(Qt.UserRole)
            if ci is not None and 0 <= ci < len(self.cached_images):
                result.append(self.cached_images[ci])
        return result


# ================== 批量图片对话框 ==================
class BatchImageDialog(QDialog):
    def __init__(self, cached_images, rows, cols, parent=None, default_sheet=None):
        super().__init__(parent)
        self.setWindowTitle("批量图片配置")
        self.setMinimumSize(750, 500)
        self.cached_images = cached_images
        self.default_sheet = default_sheet
        self.target_rows = rows
        self.target_cols = cols
        self.image_list = []

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"目标区域：{rows} 行 × {cols} 列，共需 {rows*cols} 张图片"))

        source_group = QGroupBox("图片来源")
        source_layout = QVBoxLayout(source_group)
        self.radio_internal = QRadioButton("从数据源提取")
        self.radio_file = QRadioButton("从外部文件选择")
        self.radio_internal.setChecked(True)
        source_layout.addWidget(self.radio_internal)
        source_layout.addWidget(self.radio_file)
        layout.addWidget(source_group)

        self.image_table = QTableWidget()
        self.image_table.setColumnCount(2)
        self.image_table.setHorizontalHeaderLabels(["序号", "来源"])
        self.image_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.image_table)

        btn_layout = QHBoxLayout()
        self.btn_add = QPushButton("添加图片")
        self.btn_add.clicked.connect(self.add_images)
        btn_layout.addWidget(self.btn_add)
        self.btn_clear = QPushButton("清空列表")
        self.btn_clear.clicked.connect(self.clear_images)
        btn_layout.addWidget(self.btn_clear)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def add_images(self):
        if self.radio_internal.isChecked():
            dlg = InternalImageSelectDialog(self.cached_images, self, default_sheet=self.default_sheet)
            if dlg.exec_():
                selected = dlg.get_selected_images()
                for item in selected:
                    self.image_list.append(('internal', item))
                    row = self.image_table.rowCount()
                    self.image_table.insertRow(row)
                    self.image_table.setItem(row, 0, QTableWidgetItem(str(row+1)))
                    self.image_table.setItem(row, 1, QTableWidgetItem(item[2]))
        else:
            files, _ = QFileDialog.getOpenFileNames(self, "选择图片文件", "",
                                                    "图片文件 (*.png *.jpg *.jpeg *.bmp *.gif);;所有文件 (*)")
            for f in files:
                self.image_list.append(('file', f))
                row = self.image_table.rowCount()
                self.image_table.insertRow(row)
                self.image_table.setItem(row, 0, QTableWidgetItem(str(row+1)))
                self.image_table.setItem(row, 1, QTableWidgetItem(os.path.basename(f)))

    def clear_images(self):
        self.image_list.clear()
        self.image_table.setRowCount(0)

    def get_image_sequence(self):
        return self.image_list
