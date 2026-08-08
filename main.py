import sys
import os
import multiprocessing

# 必须在任何 paddleocr/paddlex import 之前设置：
# Windows CPU 上 PaddlePaddle 3.3.x 的 PIR→oneDNN 回归会让 PP-OCRv6 推理直接崩溃；
# OpenCV 与 Paddle 的 OpenMP 重复库冲突会在 predict 时静默崩溃。
os.environ.setdefault('PADDLE_PDX_ENABLE_MKLDNN_BYDEFAULT', 'false')
os.environ.setdefault('KMP_DUPLICATE_LIB_OK', 'TRUE')

import json
import io
import re
import datetime
import time
import traceback
import xml.etree.ElementTree as ET

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QLabel,
    QFileDialog, QMessageBox, QMenu, QAction, QListWidget,
    QListWidgetItem, QAbstractItemView, QSplitter, QSplashScreen, QHeaderView,
    QDialog, QProgressDialog, QInputDialog, QLineEdit, QPlainTextEdit,
)
from PyQt5.QtGui import QColor, QFont, QIcon, QPixmap, QPainter, QPalette
from PyQt5.QtCore import QEasingCurve, QPropertyAnimation, QTimer, Qt
import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
from openpyxl.styles.colors import COLOR_INDEX
from constants import (
    DEFAULT_ROW_HEIGHT_PTS,
    ADMIN_USER_ID, ADMIN_AUTH_PASSWORD,
    MAX_PREVIEW_ROWS, MAX_PREVIEW_COLS,
)
from user_auth import load_authorized_ids, save_authorized_ids, parse_id_input
from safe_eval import _check_transform_expr
from dialogs import (
    ImageSetupDialog, TransformDialog, ArchiveConfigDialog, JMPConfigDialog,
    SourceSelectDialog, InternalImageSelectDialog, BatchImageDialog,
    VersionFinderDialog,
)
from ocr_dialog import OCRSetupDialog
from ocr_engine import ocr_available
from mapping_operations import MappingOperations
from version_finder import suggest_files
from table_zoom import TableZoomMixin
from utils import (
    column_width_chars, apply_uniform_sizes, normalize_mappings,
    make_image_thumbnail,
)
from ui_theme import (
    APPLE_QSS, CARD_PAD, C_SUB, GAP_SECTION, WINDOW_MIN_SIZE,
    window_target_size,
)
from user_guide import show_user_guide


def resource_path(relative_path):
    """获取资源文件的绝对路径，支持开发环境和 PyInstaller 打包后的环境"""
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller 打包后，资源文件在临时目录中
        return os.path.join(sys._MEIPASS, relative_path)
    # 开发环境，直接使用当前目录
    return os.path.join(os.path.abspath("."), relative_path)
# ================== 全局异常钩子 ==================
def global_exception_hook(exctype, value, tb):
    error_msg = ''.join(traceback.format_exception(exctype, value, tb))
    print(error_msg)
    # 无 QApplication 时（如导入期错误）直接退出，避免弹窗卡死
    if QApplication.instance() is None:
        sys.exit(1)
    msg = QMessageBox(QMessageBox.Critical, "程序错误", f"发生未处理异常：\n{value}")
    msg.setDetailedText(error_msg)
    msg.exec_()
    sys.exit(1)

sys.excepthook = global_exception_hook


class MainWindow(QMainWindow, MappingOperations, TableZoomMixin):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("自动M/PBO报告制作软件")
        # Apple 风格：主窗口按屏幕可用区域 80% 动态计算，最小 900×620
        w, h = window_target_size()
        self.resize(w, h)
        self.setMinimumSize(*WINDOW_MIN_SIZE)
        self.setStyleSheet(APPLE_QSS)

        self.template_wb = None
        self.source_wb = None
        self.template_path = None
        self.source_path = None
        self.current_sheet_name = None
        self.mappings = []
        self.current_selection = None
        self.cell_edits = []
        self._image_streams = []
        self.cached_images = []
        self._fill_warnings = []
        self._theme_owner = None
        self._theme_map = {}

        self.init_ui()

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(16, 12, 16, 12)
        main_layout.setSpacing(GAP_SECTION)

        # Sheet面板
        sheet_panel = QWidget()
        sheet_panel.setProperty("card", True)
        sheet_layout = QVBoxLayout(sheet_panel)
        sheet_layout.setContentsMargins(CARD_PAD, CARD_PAD, CARD_PAD, CARD_PAD)
        sheet_layout.setSpacing(6)
        sheet_label = QLabel("报告Sheet列表")
        sheet_label.setProperty("heading", True)
        sheet_layout.addWidget(sheet_label)
        self.sheet_list = QListWidget()
        self.sheet_list.itemClicked.connect(self.on_sheet_clicked)
        sheet_layout.addWidget(self.sheet_list)

        # 中间
        center_widget = QWidget()
        center_widget.setProperty("card", True)
        center_layout = QVBoxLayout(center_widget)
        center_layout.setContentsMargins(CARD_PAD, CARD_PAD, CARD_PAD, CARD_PAD)
        center_layout.setSpacing(6)
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(6)
        self.btn_open_template = QPushButton("打开报告文件")
        self.btn_open_template.setProperty("primary", True)
        self.btn_open_template.clicked.connect(self.open_template)
        btn_layout.addWidget(self.btn_open_template)
        self.btn_open_source = QPushButton("打开IPQC数据源文件")
        self.btn_open_source.setProperty("secondary", True)
        self.btn_open_source.clicked.connect(self.open_source)
        btn_layout.addWidget(self.btn_open_source)
        btn_layout.addStretch()
        guide_btn = QPushButton("使用说明")
        guide_btn.setProperty("link", True)
        guide_btn.setToolTip("查看软件使用说明")
        guide_btn.clicked.connect(self.show_user_guide)
        btn_layout.addWidget(guide_btn)
        center_layout.addLayout(btn_layout)
        self.table = QTableWidget()
        self.table.setSelectionMode(QAbstractItemView.ContiguousSelection)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        # 强制白色底，避免系统深色模式下无填充单元格显示为黑色
        table_palette = self.table.palette()
        table_palette.setColor(QPalette.Base, QColor("#FFFFFF"))
        table_palette.setColor(QPalette.AlternateBase, QColor("#F7F7F7"))
        table_palette.setColor(QPalette.Window, QColor("#FFFFFF"))
        self.table.setPalette(table_palette)
        self.enable_table_zoom(self.table)
        center_layout.addWidget(self.table)

        # 右侧
        right_widget = QWidget()
        right_widget.setProperty("card", True)
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(CARD_PAD, CARD_PAD, CARD_PAD, CARD_PAD)
        right_layout.setSpacing(6)
        mapping_label = QLabel("当前Sheet映射列表")
        mapping_label.setProperty("heading", True)
        right_layout.addWidget(mapping_label)
        self.mapping_list = QListWidget()
        self.mapping_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.mapping_list.customContextMenuRequested.connect(self.show_mapping_context_menu)
        self.mapping_list.setToolTip("右键可删除或清空映射")
        right_layout.addWidget(self.mapping_list)

        self.btn_confirm_mapping = QPushButton("确认映射")
        self.btn_confirm_mapping.clicked.connect(self.confirm_mapping)
        right_layout.addWidget(self.btn_confirm_mapping)

        self.btn_output_report = QPushButton("输出报告")
        self.btn_output_report.clicked.connect(self.output_report)
        right_layout.addWidget(self.btn_output_report)

        self.btn_save_config = QPushButton("保存配置")
        self.btn_save_config.clicked.connect(self.save_config)
        right_layout.addWidget(self.btn_save_config)

        self.btn_import_config = QPushButton("导入我的配置")
        self.btn_import_config.clicked.connect(self.import_config)
        right_layout.addWidget(self.btn_import_config)

        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(sheet_panel)
        splitter.addWidget(center_widget)
        splitter.addWidget(right_widget)
        splitter.setSizes([220, 700, 280])
        main_layout.addWidget(splitter)

        # 底部版权信息
        sb = self.statusBar()
        sb.showMessage("就绪")
        sb_label = QLabel("Copyright © 2026 ABU NPD EOL")
        sb_label.setStyleSheet(f"color: {C_SUB}; font-size: 11px;")
        sb.addPermanentWidget(sb_label)

    def show_user_guide(self):
        """右上角「使用说明」：应用内帮助对话框。"""
        show_user_guide(self)

    # ==================== 文件加载 ====================
    def open_template(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择模板", "", "Excel文件 (*.xlsx)")
        if not path:
            return
        self.template_path = path
        try:
            self.template_wb = openpyxl.load_workbook(path)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法打开模板文件：{e}")
            return
        self._snapshot_template_images()

        # 配置不再随报告自动加载，由用户点击“导入我的配置”手动导入
        self.mappings = []

        self.sheet_list.clear()
        for name in self.template_wb.sheetnames:
            item = QListWidgetItem(name)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            self.sheet_list.addItem(item)

        if self.sheet_list.count() > 0:
            self.sheet_list.setCurrentRow(0)
            self.current_sheet_name = self.template_wb.sheetnames[0]
            self.display_sheet(self.template_wb[self.current_sheet_name])

    def open_source(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择数据源", "", "Excel文件 (*.xlsx)")
        if not path:
            return
        self.source_path = path
        busy = self._make_busy_progress("正在打开 IPQC 数据源，请稍等…")
        try:
            self.source_wb = openpyxl.load_workbook(path)
        except Exception as e:
            self._close_progress(busy)
            QMessageBox.critical(self, "错误", f"无法打开数据源文件：{e}")
            return
        self._close_progress(busy)
        try:
            self._cache_all_images()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"无法打开数据源文件：{e}")
            return
        QMessageBox.information(self, "提示", f"数据源文件已加载：{os.path.basename(path)}")

    def _cache_all_images(self):
        self.cached_images = []
        total = sum(len(ws._images) for ws in self.source_wb.worksheets)
        prog = self._make_progress("正在读取数据源图片，请稍等…", total)
        done = 0
        failed = []
        for sheet_name in self.source_wb.sheetnames:
            ws = self.source_wb[sheet_name]
            for idx, img in enumerate(ws._images):
                if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                    from_cell = img.anchor._from
                    row = from_cell.row + 1
                    col = from_cell.col + 1
                    pos = f"{get_column_letter(col)}{row}"
                else:
                    pos = "未知"
                try:
                    img_data = img._data()
                except AttributeError:
                    try:
                        img_data = img.ref.read()
                    except Exception as e:
                        failed.append(f"{sheet_name}!{pos}（{e}）")
                        done += 1
                        self._update_progress(prog, done, total)
                        continue
                width = getattr(img, 'width', '')
                height = getattr(img, 'height', '')
                # 同时生成 JPEG 小缩略图供选择/预览用，避免预览时全图解码卡顿；
                # 输出报告仍用原始 img_data，不受影响
                thumb = make_image_thumbnail(img_data)
                self.cached_images.append(
                    (sheet_name, idx, pos, img_data, width, height, thumb))
                done += 1
                self._update_progress(prog, done, total)
        self._close_progress(prog)
        if failed:
            detail = "\n".join(failed[:10])
            more = len(failed) - 10
            suffix = f"\n…另有 {more} 处" if more > 0 else ""
            if total > 0 and len(failed) == total:
                QMessageBox.warning(
                    self, "图片读取异常",
                    f"数据源 {len(failed)} 张图片全部读取失败，"
                    "文件可能已损坏，请确认后重新打开：\n\n" + detail + suffix)
            else:
                QMessageBox.warning(
                    self, "部分图片读取失败",
                    f"有 {len(failed)}/{total} 张图片读取失败，"
                    "对应图片映射将无法输出，请检查数据源：\n\n" + detail + suffix)

    def _make_progress(self, label, total):
        """创建模态进度条弹窗（不可取消；立即显示）"""
        if total <= 0:
            return None
        dlg = QProgressDialog(label, None, 0, total, self)
        dlg.setWindowTitle("处理中")
        dlg.setWindowModality(Qt.WindowModal)
        dlg.setCancelButton(None)
        dlg.setMinimumDuration(0)
        dlg.setValue(0)
        dlg.show()
        QApplication.processEvents()
        return dlg

    def _make_busy_progress(self, label):
        """创建不确定进度的等待弹窗（用于保存等无法计数的阶段）"""
        dlg = QProgressDialog(label, None, 0, 0, self)
        dlg.setWindowTitle("处理中")
        dlg.setWindowModality(Qt.WindowModal)
        dlg.setCancelButton(None)
        dlg.setMinimumDuration(0)
        dlg.setValue(0)
        dlg.show()
        QApplication.processEvents()
        return dlg

    @staticmethod
    def _update_progress(dlg, value, total):
        if dlg is None or total <= 0:
            return
        dlg.setValue(min(value, total))
        QApplication.processEvents()

    @staticmethod
    def _close_progress(dlg):
        if dlg is not None:
            dlg.close()
            QApplication.processEvents()

    @staticmethod
    def _read_image_ref(ref):
        """读取图片字节而不关闭其文件对象"""
        try:
            if isinstance(ref, io.BytesIO):
                return ref.getvalue()
            ref.seek(0)
            data = ref.read()
            ref.seek(0)
            return data
        except Exception:
            return None

    def _snapshot_template_images(self):
        """记录模板内嵌图片字节，供保存前重建 ref。
        openpyxl 保存图片后会 close 掉 BytesIO ref，第二次保存会报
        'I/O operation on closed file'，因此保存前需用快照重建。"""
        for name in self.template_wb.sheetnames:
            for img in self.template_wb[name]._images:
                img._saved_bytes = self._read_image_ref(img.ref)

    def _refresh_image_refs(self):
        """保存前用字节快照重建所有图片 ref"""
        for name in self.template_wb.sheetnames:
            for img in self.template_wb[name]._images:
                data = getattr(img, '_saved_bytes', None)
                if data:
                    img.ref = io.BytesIO(data)

    # ==================== Sheet渲染 ====================
    def on_sheet_clicked(self, item):
        self.current_sheet_name = item.text()
        self.display_sheet(self.template_wb[self.current_sheet_name])
        self.refresh_mapping_list()

    # 预览不允许出现黑色单元格：深/黑色填充（如主题 dk1=#000000、
    # indexed 8）一律替换为浅灰，保证预览干净可读（输出文件不受影响）
    PREVIEW_DARK_FILL_THRESHOLD = 0.5
    PREVIEW_DARK_FILL_REPLACEMENT = "#D9D9D9"

    def _cell_fill_color(self, cell, wb):
        """解析单元格填充色；无填充/无法解析返回 None。
        深/黑色填充在预览中替换为浅灰，不允许出现黑色单元格。"""
        fill = cell.fill
        if not fill or not fill.patternType:
            return None
        qc = self._resolve_color(fill.fgColor, wb)
        if qc is None:
            return None
        if self._luminance(qc) < self.PREVIEW_DARK_FILL_THRESHOLD:
            return QColor(self.PREVIEW_DARK_FILL_REPLACEMENT)
        return qc

    def _cell_font_color(self, cell, wb):
        """解析单元格字体色。
        Excel/WPS 把“自动”文本色序列化为 theme=0/1、indexed=8/9/64/65，
        实际显示时按背景亮度自适应：深底显示白字、浅底/无底显示黑字，
        因此这里单独处理；显式 rgb 色则原样使用。未设置颜色等同自动色。"""
        color = cell.font.color if cell.font else None
        is_auto = color is None
        if color is not None:
            ctype = getattr(color, 'type', None)
            try:
                if ctype == 'theme':
                    is_auto = color.theme in (0, 1)
                elif ctype == 'indexed':
                    # 8=自动文字(黑), 9=自动背景(白), 64/65=自动前景/背景
                    is_auto = color.indexed in (8, 9, 64, 65)
                elif ctype == 'auto':
                    is_auto = True
            except Exception:
                pass
        if is_auto:
            bg = self._cell_fill_color(cell, wb)
            if bg is not None and self._luminance(bg) < 0.5:
                return QColor("#FFFFFF")
            return QColor("#000000")
        return self._resolve_color(color, wb)

    @staticmethod
    def _luminance(qcolor):
        """按人眼感知亮度计算颜色亮度（0~1）"""
        return (0.299 * qcolor.red() + 0.587 * qcolor.green()
                + 0.114 * qcolor.blue()) / 255.0

    @staticmethod
    def _apply_tint(hexcolor, tint):
        """对主题色应用 tint（OOXML 规范）：
        负值向黑加深、正值向白变浅。"""
        if not tint:
            return hexcolor
        h = hexcolor.lstrip('#')
        try:
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        except (ValueError, IndexError):
            return hexcolor
        if tint < 0:
            factor = 1.0 + tint
            r, g, b = r * factor, g * factor, b * factor
        else:
            r = r * (1.0 - tint) + 255.0 * tint
            g = g * (1.0 - tint) + 255.0 * tint
            b = b * (1.0 - tint) + 255.0 * tint
        return "#%02X%02X%02X" % (
            max(0, min(255, int(round(r)))),
            max(0, min(255, int(round(g)))),
            max(0, min(255, int(round(b)))),
        )

    def _resolve_color(self, color, wb):
        """把 openpyxl 颜色对象解析为 QColor；rgb/indexed/theme 三类均支持"""
        if color is None:
            return None
        ctype = getattr(color, 'type', None)
        try:
            if ctype == 'rgb':
                rgb = color.rgb
                if isinstance(rgb, str) and len(rgb) >= 6:
                    qc = QColor(f"#{rgb[-6:]}")
                    return qc if qc.isValid() else None
            elif ctype == 'indexed':
                idx = color.indexed
                if isinstance(idx, int) and 0 <= idx < len(COLOR_INDEX):
                    entry = COLOR_INDEX[idx]
                    if entry:
                        qc = QColor(f"#{entry[-6:]}")
                        return qc if qc.isValid() else None
            elif ctype == 'theme':
                hexval = self._theme_color(wb, color.theme)
                if hexval:
                    tint = getattr(color, 'tint', None)
                    qc = QColor(self._apply_tint(hexval, tint) if tint else hexval)
                    return qc if qc.isValid() else None
        except Exception:
            pass
        return None

    def _theme_color(self, wb, idx):
        """按主题色索引返回颜色（0=dk1,1=lt1,2=dk2,3=lt2,4-9=accent1-6,...）"""
        if self._theme_owner is not wb:
            self._theme_owner = wb
            self._theme_map = self._parse_theme_colors(wb)
        return self._theme_map.get(idx)

    @staticmethod
    def _parse_theme_colors(wb):
        """解析工作簿主题 XML 中的 clrScheme 颜色"""
        theme_bytes = getattr(wb, 'loaded_theme', None)
        if not theme_bytes:
            return {}
        ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
        try:
            root = ET.fromstring(theme_bytes)
            scheme = root.find('.//a:clrScheme', ns)
            if scheme is None:
                return {}
            order = ['dk1', 'lt1', 'dk2', 'lt2',
                     'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6',
                     'hlink', 'folHlink']
            result = {}
            for i, name in enumerate(order):
                el = scheme.find(f'a:{name}', ns)
                if el is None:
                    continue
                srgb = el.find('a:srgbClr', ns)
                if srgb is not None and 'val' in srgb.attrib:
                    result[i] = f"#{srgb.attrib['val']}"
                    continue
                sysclr = el.find('a:sysClr', ns)
                if sysclr is not None and 'lastClr' in sysclr.attrib:
                    result[i] = f"#{sysclr.attrib['lastClr']}"
            return result
        except Exception:
            return {}

    def display_sheet(self, ws):
        self.table.clear()
        self.zoom_begin()
        self.table.clearSpans()
        real_max_row = ws.max_row
        real_max_col = ws.max_column
        if real_max_row > 200:
            for r in range(real_max_row, 0, -1):
                if any(cell.value is not None for cell in ws[r]):
                    real_max_row = r
                    break
            real_max_row += 2
        if real_max_col > 30:
            for c in range(real_max_col, 0, -1):
                if any(ws.cell(row=r, column=c).value is not None for r in range(1, real_max_row+1)):
                    real_max_col = c
                    break
            real_max_col += 2

        max_rows = min(real_max_row, MAX_PREVIEW_ROWS)
        max_cols = min(real_max_col, MAX_PREVIEW_COLS)
        self.table.setRowCount(max_rows)
        self.table.setColumnCount(max_cols)

        try:
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                if min_row > max_rows or min_col > max_cols:
                    continue
                span_rows = min(max_row, max_rows) - min_row + 1
                span_cols = min(max_col, max_cols) - min_col + 1
                self.table.setSpan(min_row-1, min_col-1, span_rows, span_cols)

            wb = ws.parent
            # 稀疏遍历：只处理工作簿中真实存在的单元格，
            # 避免 iter_rows 为超大表物化大量空单元格导致卡顿
            for (row_idx, col_idx), cell in ws._cells.items():
                if row_idx > max_rows or col_idx > max_cols:
                    continue
                if cell.value is None:
                    # 空单元格只要渲染填充色（字体色无意义），无填充则跳过
                    bg = self._cell_fill_color(cell, wb)
                    if bg is None:
                        continue
                    item = QTableWidgetItem("")
                    item.setBackground(bg)
                    self.table.setItem(row_idx - 1, col_idx - 1, item)
                    continue
                item = QTableWidgetItem(str(cell.value))
                base_size = 10
                if cell.font:
                    font = QFont()
                    font.setFamily(cell.font.name or 'Arial')
                    size = cell.font.size or 10
                    if size is not None:
                        font.setPointSize(int(size))
                        base_size = int(size)
                    font.setBold(cell.font.bold)
                    item.setFont(font)
                self.zoom_track_item(row_idx - 1, col_idx - 1, base_size)
                bg = self._cell_fill_color(cell, wb)
                if bg is not None:
                    item.setBackground(bg)
                fg = self._cell_font_color(cell, wb)
                if fg is not None:
                    item.setForeground(fg)
                self.table.setItem(row_idx - 1, col_idx - 1, item)
        except Exception as e:
            QMessageBox.warning(self, "渲染警告", f"表格渲染出现异常：{e}")

        apply_uniform_sizes(self.table, ws, max_cols, max_rows, zoom=self)

    # ==================== 右键菜单 ====================
    def show_context_menu(self, pos):
        menu = QMenu()

        # 右键单元格 -> 修改内容（可作用于空白单元格，用于新增内容）
        row = self.table.rowAt(pos.y()) + 1
        col = self.table.columnAt(pos.x()) + 1
        if 1 <= row <= self.table.rowCount() and 1 <= col <= self.table.columnCount():
            ws = self.template_wb[self.current_sheet_name]
            cell_ref = f"{get_column_letter(col)}{row}"
            menu.addAction(QAction(f"修改单元格内容（{cell_ref}）", self,
                                   triggered=lambda: self.edit_cell_value(ws, row, col)))
            menu.addSeparator()

        indexes = self.table.selectedIndexes()
        if indexes:
            top = min(idx.row() for idx in indexes) + 1
            left = min(idx.column() for idx in indexes) + 1
            bottom = max(idx.row() for idx in indexes) + 1
            right = max(idx.column() for idx in indexes) + 1
            self.current_selection = (top, left, bottom, right)
            info_action = QAction(
                f"选中区域: {get_column_letter(left)}{top}:{get_column_letter(right)}{bottom}  ({bottom-top+1}行×{right-left+1}列)",
                self
            )
            info_action.setEnabled(False)
            menu.addAction(info_action)
            menu.addSeparator()
            menu.addAction(QAction("设为数据填充区", self, triggered=self.add_data_mapping))
            menu.addAction(QAction("设为图片区域", self, triggered=self.add_image_mapping))
            menu.addAction(QAction("设为归档区域（右移）", self, triggered=self.add_archive_mapping))
            menu.addAction(QAction("设为JMP数据区", self, triggered=self.add_jmp_mapping))
            # OCR 切片量测（需要 PaddleOCR）
            ocr_action = QAction("OCR识别（切片量测）", self, triggered=self.add_ocr_mapping)
            if not ocr_available():
                ocr_action.setEnabled(False)
                ocr_action.setToolTip("PaddleOCR 未安装，请运行 pip install paddlepaddle paddleocr")
            menu.addAction(ocr_action)
        if self.template_wb and self._version_header_cells(
                self.template_wb[self.current_sheet_name]):
            menu.addSeparator()
            menu.addAction(QAction("查找版本号…", self, triggered=self.find_version_numbers))

        if not menu.actions():
            return
        menu.exec_(self.table.viewport().mapToGlobal(pos))

    def edit_cell_value(self, ws, row, col):
        """右键修改单元格内容，同步到模板工作簿（输出报告时生效）"""
        item = self.table.item(row - 1, col - 1)
        current = item.text() if item else ""
        new_text, ok = QInputDialog.getText(
            self, "修改单元格内容",
            f"单元格 {get_column_letter(col)}{row} 的新内容（留空=清空，=开头=公式）：",
            QLineEdit.Normal, current)
        if not ok:
            return
        value = self._parse_cell_input(new_text)
        ws.cell(row=row, column=col).value = value
        self.cell_edits.append([self.current_sheet_name, row, col, value])
        if item is None:
            item = QTableWidgetItem()
            self.table.setItem(row - 1, col - 1, item)
        item.setText("" if value is None else str(value))

    @staticmethod
    def _parse_cell_input(text):
        """把输入文本转为单元格值：数字->数值，=开头->公式，空->None，其余->文本"""
        text = text.strip()
        if text == "":
            return None
        if text.startswith("="):
            return text
        try:
            return int(text)
        except ValueError:
            pass
        try:
            return float(text)
        except ValueError:
            return text

    # ==================== 映射添加 ====================
    def add_data_mapping(self):
        if not self.current_selection or not self.source_wb:
            QMessageBox.warning(self, "提示", "请先打开数据源文件并选中目标区域")
            return
        dlg = SourceSelectDialog(self.source_wb, self, default_sheet=self.current_sheet_name)
        if dlg.exec_():
            src_sheet, src_range = dlg.get_selection()
            if not src_range:
                QMessageBox.warning(self, "提示", "源区域不能为空")
                return

            trans_dlg = TransformDialog(self)
            if not trans_dlg.exec_():
                return
            trans_type, trans_expr = trans_dlg.get_transform()
            clear_target = trans_dlg.get_clear_target()

            if trans_type == 'custom' and trans_expr.strip():
                err = _check_transform_expr(trans_expr.strip())
                if err:
                    QMessageBox.warning(self, "表达式无效",
                        f"自定义表达式无法安全解析：{trans_expr}\n{err}\n\n"
                        "支持：x 与四则运算、比较、abs/round/min/max/int/float/str/len/sum。")
                    return

            # 防止数据填充目标与归档区块重叠（归档以磁盘原始数据为基准，会覆盖填充结果）
            for m in self.mappings:
                if (m.get('target_sheet') != self.current_sheet_name
                        or m.get('type') != 'archive_shift_right'):
                    continue
                if self._archive_blocks_overlap(self.current_selection, m['block_range']):
                    QMessageBox.warning(self, "区域冲突",
                        f"数据填充区域与归档区块重叠，无法添加：\n"
                        f"填充区域：{self._block_range_str(self.current_selection)}\n"
                        f"归档区块：{self._block_range_str(m['block_range'])}\n\n"
                        "归档会以磁盘原始数据为基准移位，重叠的填充结果会被覆盖。")
                    return

            self.mappings.append({
                'type': 'data',
                'target_sheet': self.current_sheet_name,
                'target_range': self.current_selection,
                'source_sheet': src_sheet,
                'source_range': src_range,
                'transform': trans_type,
                'transform_expr': trans_expr,
                'clear_target': clear_target
            })
            self.refresh_mapping_list()

    def add_image_mapping(self):
        if not self.current_selection:
            QMessageBox.warning(self, "提示", "请先在模板中选中目标区域")
            return

        t_min_row, t_min_col, _, _ = self.current_selection
        ws = self.template_wb[self.current_sheet_name]
        default_w, default_h = self._cell_default_size(ws, t_min_row, t_min_col)
        dlg = ImageSetupDialog(self, default_col_width=default_w,
                               default_row_height=default_h)
        if not dlg.exec_():
            return
        col_width_chars, row_height_pts, rotation, w_scale, h_scale, alignment = dlg.get_values()

        _, _, t_max_row, t_max_col = self.current_selection
        rows = t_max_row - t_min_row + 1
        cols = t_max_col - t_min_col + 1
        if rows * cols == 1:
            self._add_single_image_mapping(col_width_chars, row_height_pts, rotation, w_scale, h_scale, alignment)
        else:
            self._add_batch_image_mapping(rows, cols, col_width_chars, row_height_pts, rotation, w_scale, h_scale, alignment)

    def _cell_default_size(self, ws, row, col):
        """锚点单元格（或所在合并区域）的默认图片尺寸：列宽(字符)、行高(磅)。
        合并单元格按整个合并区域求和，未合并则取单格大小。"""
        col_span = row_span = 1
        for mr in ws.merged_cells.ranges:
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(mr))
            if m_min_row <= row <= m_max_row and m_min_col <= col <= m_max_col:
                col_span = m_max_col - m_min_col + 1
                row_span = m_max_row - m_min_row + 1
                break
        width_chars = sum(
            column_width_chars(ws, c)
            for c in range(col, col + col_span)
        )
        height_pts = 0.0
        for r in range(row, row + row_span):
            if r in ws.row_dimensions and ws.row_dimensions[r].height:
                height_pts += ws.row_dimensions[r].height
            else:
                height_pts += ws.sheet_format.defaultRowHeight or DEFAULT_ROW_HEIGHT_PTS
        return round(width_chars, 1), round(height_pts, 1)

    def _add_single_image_mapping(self, col_width_chars, row_height_pts, rotation, w_scale, h_scale, alignment='left'):
        t_row, t_col, _, _ = self.current_selection
        anchor = f"{get_column_letter(t_col)}{t_row}"

        if not self.source_wb:
            QMessageBox.information(self, "提示", "未打开数据源文件，只能从外部文件选择图片")
            img_path, _ = QFileDialog.getOpenFileName(self, "选择图片文件", "",
                                                       "图片文件 (*.png *.jpg *.jpeg *.bmp *.gif);;所有文件 (*)")
            if img_path:
                mapping = {
                    'type': 'image',
                    'target_sheet': self.current_sheet_name,
                    'anchor_cell': anchor,
                    'image_path': img_path,
                    'col_width_chars': col_width_chars,
                    'row_height_pts': row_height_pts,
                    'rotation': rotation,
                    'width_scale': w_scale,
                    'height_scale': h_scale,
                    'alignment': alignment,
                }
                self.mappings.append(mapping)
                self.refresh_mapping_list()
            return

        if not self.cached_images:
            reply = QMessageBox.question(self, "图片来源", "数据源中没有图片，是否从外部文件选择？", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                img_path, _ = QFileDialog.getOpenFileName(self, "选择图片文件", "",
                                                           "图片文件 (*.png *.jpg *.jpeg *.bmp *.gif);;所有文件 (*)")
                if img_path:
                    mapping = {
                        'type': 'image',
                        'target_sheet': self.current_sheet_name,
                        'anchor_cell': anchor,
                        'image_path': img_path,
                        'col_width_chars': col_width_chars,
                        'row_height_pts': row_height_pts,
                        'rotation': rotation,
                        'width_scale': w_scale,
                        'height_scale': h_scale,
                        'alignment': alignment,
                    }
                    self.mappings.append(mapping)
                    self.refresh_mapping_list()
            return

        dlg = InternalImageSelectDialog(self.cached_images, self, default_sheet=self.current_sheet_name)
        if dlg.exec_():
            selected = dlg.get_selected_images()
            if selected:
                img_info = selected[0]
                mapping = {
                    'type': 'image',
                    'target_sheet': self.current_sheet_name,
                    'anchor_cell': anchor,
                    'image_bytes': img_info[3],
                    'image_ref': [img_info[0], img_info[1]],
                    'image_src_sheet': img_info[0],
                    'image_src_pos': img_info[2],
                    'orig_width': img_info[4],
                    'orig_height': img_info[5],
                    'col_width_chars': col_width_chars,
                    'row_height_pts': row_height_pts,
                    'rotation': rotation,
                    'width_scale': w_scale,
                    'height_scale': h_scale,
                    'alignment': alignment,
                }
                self.mappings.append(mapping)
                self.refresh_mapping_list()

    def _add_batch_image_mapping(self, rows, cols, col_width_chars, row_height_pts, rotation, w_scale, h_scale, alignment='left'):
        if not self.source_wb:
            QMessageBox.warning(self, "提示", "请先打开数据源文件")
            return
        if not self.cached_images:
            QMessageBox.warning(self, "提示", "数据源中没有图片，无法批量添加")
            return
        dlg = BatchImageDialog(self.cached_images, rows, cols, self,
                               default_sheet=self.current_sheet_name)
        if dlg.exec_() != QDialog.Accepted:
            return
        images = dlg.get_image_sequence()
        if not images:
            QMessageBox.warning(self, "图片不足", "未选择任何图片，无法添加批量图片映射")
            return
        need = rows * cols
        if len(images) > need:
            QMessageBox.warning(self, "图片超出目标区域",
                f"目标区域共 {need} 个单元格，选择了 {len(images)} 张图片，"
                "仅保留前 " + str(need) + " 张。")
            images = images[:need]
        t_min_row, t_min_col, _, _ = self.current_selection
        _, _, t_max_row, t_max_col = self.current_selection
        # PBO→MBO 时图片数少于模板区域：整区清空旧图片，新图按顺序从左上角填入，
        # 未填位置保持空白，避免残留 PBO 图片
        clear_region = (t_min_row, t_min_col, t_max_row, t_max_col)
        count = 0
        for r in range(rows):
            for c in range(cols):
                if count >= len(images):
                    break
                anchor = f"{get_column_letter(t_min_col + c)}{t_min_row + r}"
                img_type, img_data = images[count]
                count += 1
                base_mapping = {
                    'type': 'image',
                    'target_sheet': self.current_sheet_name,
                    'anchor_cell': anchor,
                    'clear_region': clear_region,
                    'col_width_chars': col_width_chars,
                    'row_height_pts': row_height_pts,
                    'rotation': rotation,
                    'width_scale': w_scale,
                    'height_scale': h_scale,
                    'alignment': alignment,
                }
                if img_type == 'file':
                    base_mapping['image_path'] = img_data
                else:
                    base_mapping['image_bytes'] = img_data[3]
                    base_mapping['image_ref'] = [img_data[0], img_data[1]]
                    base_mapping['image_src_sheet'] = img_data[0]
                    base_mapping['image_src_pos'] = img_data[2]
                    base_mapping['orig_width'] = img_data[4]
                    base_mapping['orig_height'] = img_data[5]
                self.mappings.append(base_mapping)
        self.refresh_mapping_list()

    # ==================== 版本号查找 ====================
    VERSION_HEADERS = {
        'process_control': 'Process control rev.',
        'ers': 'ERS rev.',
        'vsr': 'VSR rev.',
        'mco': 'MCO rev.',
    }

    def _version_header_cells(self, ws):
        """在当前表定位版本号表头及其下一行的值单元格"""
        found = {}
        for key, header in self.VERSION_HEADERS.items():
            for row in ws.iter_rows():
                for cell in row:
                    if (isinstance(cell.value, str)
                            and cell.value.strip().lower() == header.lower()):
                        found[key] = (cell.row + 1, cell.column)
                        break
                if key in found:
                    break
        return found

    def _suggest_version_files(self):
        """从模板所在目录自动推荐 ERS/VSR/MCO 档案。
        只有 CLO 专案推荐 099-55402 系 ERS，其他专案优先推荐 BUF 模式。"""
        base = os.path.dirname(self.template_path) if self.template_path else ""
        keyword = 'BUF'
        if self.template_path:
            m = re.match(r'^([A-Za-z]+)', os.path.basename(self.template_path))
            if m and m.group(1).upper() == 'CLO':
                keyword = 'CLO'
        suggestions = suggest_files(base, keyword)
        suggestions['process_control'] = suggestions.get('ers')
        return suggestions

    def find_version_numbers(self):
        ws = self.template_wb[self.current_sheet_name]
        header_cells = self._version_header_cells(ws)
        if not header_cells:
            QMessageBox.information(self, "提示", "当前表未找到版本号字段（Process control/ERS/VSR/MCO rev.）")
            return
        keyword = 'BUF'
        if self.template_path:
            m = re.match(r'^([A-Za-z]+)', os.path.basename(self.template_path))
            if m and m.group(1).upper() == 'CLO':
                keyword = 'CLO'
        dlg = VersionFinderDialog(
            self, suggestions=self._suggest_version_files(),
            folder_keyword=keyword)
        if dlg.exec_() != QDialog.Accepted:
            return
        results = dlg.get_results()
        filled = 0
        skipped = []
        for key, (path, value) in results.items():
            if key not in header_cells:
                skipped.append(key)
                continue
            if value is None:
                skipped.append(key)
                continue
            if key == 'process_control':
                try:
                    # 数值版本号（如 3.55）保留合理精度，避免浮点二进制误差
                    value = round(float(value), 6) if re.match(r'^\d+\.?\d*$', value) else value
                except ValueError:
                    pass
            else:
                try:
                    value = int(float(value))
                except ValueError:
                    pass
            row, col = header_cells[key]
            ws.cell(row=row, column=col).value = value
            self.cell_edits.append([self.current_sheet_name, row, col, value])
            filled += 1
        self.display_sheet(ws)
        if skipped:
            QMessageBox.warning(
                self, "部分字段未更新",
                f"以下字段匹配失败，已跳过更新：\n{', '.join(skipped)}\n\n"
                "请在输出报告后，在报告中手动修改这些单元格。")
        else:
            QMessageBox.information(self, "完成", f"已写入 {filled} 个版本号字段")

    def add_archive_mapping(self):
        if not self.current_selection:
            QMessageBox.warning(self, "提示", "请先选中归档区域的首列")
            return
        t_min_row, t_min_col, t_max_row, _ = self.current_selection
        ws = self.template_wb[self.current_sheet_name]

        dlg = ArchiveConfigDialog(template_range=(t_min_row, t_min_col, t_max_row, t_min_col), parent=self)
        if dlg.exec_():
            header_rows, headers, source_col = dlg.get_selection()
            if not source_col:
                QMessageBox.warning(self, "提示", "请填写新数据的来源列（如 J）")
                return
            try:
                column_index_from_string(source_col)
            except ValueError:
                QMessageBox.warning(self, "提示", f"来源列无效：{source_col}")
                return

            # 先按实际内容收缩最大列：跳过只设格式、没有内容的远距离单元格，
            # 避免扫描范围被 max_column 撑大
            content_max_col = max(
                (col for (row, col), cell in ws._cells.items() if cell.value is not None),
                default=0,
            )
            max_col = min(ws.max_column, content_max_col)
            right_col = t_min_col
            # 只扫表头行：从锚定列向右，找到最后一个表头有内容的列（无列数上限）
            for col in range(t_min_col, max_col + 1):
                has_value = False
                for r in range(t_min_row, t_min_row + header_rows):
                    cell = ws.cell(row=r, column=col)
                    if cell.value is not None:
                        has_value = True
                        break
                if has_value:
                    right_col = col
                else:
                    break

            full_block = (t_min_row, t_min_col, t_max_row, right_col)
            block_str = f"{get_column_letter(t_min_col)}{t_min_row}:{get_column_letter(right_col)}{t_max_row}"
            reply = QMessageBox.question(self, "确认归档范围",
                f"自动检测到归档区域：\n{block_str}\n（首列 {get_column_letter(t_min_col)} → 末列 {get_column_letter(right_col)}）\n\n是否继续？",
                QMessageBox.Yes | QMessageBox.No)
            if reply != QMessageBox.Yes:
                return

            # 完全相同的归档映射已存在时不重复添加（避免操作重复触发）
            for m in self.mappings:
                if (m.get('target_sheet') == self.current_sheet_name
                        and m.get('type') == 'archive_shift_right'
                        and m.get('block_range') == full_block):
                    QMessageBox.information(self, "提示",
                        "该区块已存在相同的归档映射，无需重复添加。")
                    return

            # 防止多个归档区块重叠：重叠区块会互相覆盖数据（归档以磁盘原始数据为基准），
            # 直接拒绝添加并提示用户调整范围
            for m in self.mappings:
                if (m.get('target_sheet') != self.current_sheet_name
                        or m.get('type') != 'archive_shift_right'):
                    continue
                if self._archive_blocks_overlap(full_block, m['block_range']):
                    QMessageBox.warning(
                        self, "归档区域冲突",
                        f"该归档区块与已有归档映射重叠，无法添加：\n"
                        f"当前区块：{block_str}\n"
                        f"已有区块：{self._block_range_str(m['block_range'])}\n\n"
                        "归档区块之间必须互不重叠（行、列区间不能同时相交）。")
                    return

            # 防止归档区块与已有数据填充区域重叠（同上原因）
            for m in self.mappings:
                if (m.get('target_sheet') != self.current_sheet_name
                        or m.get('type') != 'data'):
                    continue
                if self._archive_blocks_overlap(full_block, m['target_range']):
                    QMessageBox.warning(self, "区域冲突",
                        f"归档区块与已有数据填充区域重叠，无法添加：\n"
                        f"归档区块：{block_str}\n"
                        f"填充区域：{self._block_range_str(m['target_range'])}\n\n"
                        "归档会以磁盘原始数据为基准移位，重叠的填充结果会被覆盖。")
                    return

            self.mappings.append({
                'type': 'archive_shift_right',
                'target_sheet': self.current_sheet_name,
                'block_range': full_block,
                'header_rows': header_rows,
                'new_headers': headers,
                'source_col': source_col
            })
            self.refresh_mapping_list()

    def add_jmp_mapping(self):
        if not self.current_selection:
            QMessageBox.warning(self, "提示", "请先选中JMP数据区的起始单元格（锚点）")
            return
        t_row, t_col, _, _ = self.current_selection
        anchor = f"{get_column_letter(t_col)}{t_row}"

        dlg = JMPConfigDialog(self.template_wb, anchor, self)
        if dlg.exec_():
            config = dlg.get_config()
            if not config['source_range']:
                QMessageBox.warning(self, "提示", "源区域不能为空")
                return

            mapping = {
                'type': 'jmp',
                'target_sheet': self.current_sheet_name,
                'anchor_cell': anchor,
                'header_cols': config['header_cols'],  # 列表
                'source_sheet': config['source_sheet'],
                'source_range': config['source_range'],
                'merge_columns': config['merge_columns']
            }
            self.mappings.append(mapping)
            self.refresh_mapping_list()

    # ==================== OCR 切片量测 ====================
    def add_ocr_mapping(self):
        """右键 → OCR识别（切片量测）：弹出配置对话框，创建 OCR 映射。"""
        if not ocr_available():
            QMessageBox.warning(self, "OCR 不可用",
                "PaddleOCR 未安装。请运行: pip install paddlepaddle paddleocr")
            return

        dlg = OCRSetupDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return

        mapping = dlg.get_ocr_mapping()
        mapping['target_sheet'] = self.current_sheet_name

        # 自动计算 target_cells：按选中的单元格区域展开
        img_count = dlg.get_selected_count()
        labels = dlg.get_labels()
        mode = mapping['mode']

        if mode == 'labeled' and labels:
            # 多值模式：每张图一行，每标签占一列
            cell_count = img_count * len(labels)
            if self.current_selection:
                t_min_row, t_min_col, t_max_row, t_max_col = self.current_selection
                selected_cells = (t_max_row - t_min_row + 1) * (t_max_col - t_min_col + 1)
                if selected_cells == cell_count:
                    # 完美匹配：按行展开（每张图一行，每标签一列）
                    mapping['target_cells'] = []
                    cols_per_row = len(labels)
                    for i in range(img_count):
                        row = t_min_row + i
                        cells = []
                        for j in range(cols_per_row):
                            col = t_min_col + j
                            cells.append([row, col])
                        mapping['target_cells'].append(cells)
                elif selected_cells == img_count:
                    # 只选了单列：每张图一个单元格，取第一个标签值
                    mapping['target_cells'] = []
                    for i in range(img_count):
                        row = t_min_row + i
                        mapping['target_cells'].append([row, t_min_col])
                    # 降级为单值模式
                    mapping['mode'] = 'first_number'
                else:
                    QMessageBox.warning(self, "单元格数量不匹配",
                        f"选中了 {selected_cells} 个单元格，但需要 {cell_count} 个"
                        f"（{img_count} 张图 × {len(labels)} 个标签）。\n\n"
                        "请重新选择与图像数量匹配的单元格区域。")
                    return
            else:
                QMessageBox.warning(self, "提示",
                    "请先在报告模板中选中目标单元格区域，再添加 OCR 映射。")
                return
        else:
            # 单值模式：每张图一个单元格
            if self.current_selection:
                t_min_row, t_min_col, t_max_row, t_max_col = self.current_selection
                selected_cells = (t_max_row - t_min_row + 1) * (t_max_col - t_min_col + 1)
                if selected_cells == img_count:
                    mapping['target_cells'] = []
                    for i in range(img_count):
                        row = t_min_row + i
                        for j in range(t_max_col - t_min_col + 1):
                            col = t_min_col + j
                            mapping['target_cells'].append([row, col])
                elif selected_cells == 1:
                    # 只选了一个单元格：所有图的结果填入同一格（不推荐但允许）
                    row, col = t_min_row, t_min_col
                    mapping['target_cells'] = [[row, col] for _ in range(img_count)]
                else:
                    QMessageBox.warning(self, "单元格数量不匹配",
                        f"选中了 {selected_cells} 个单元格，但有 {img_count} 张图像。\n\n"
                        "请确保选中单元格数等于图像数（单值模式）或 "
                        "图像数 × 标签数（多值模式）。")
                    return
            else:
                QMessageBox.warning(self, "提示",
                    "请先在报告模板中选中目标单元格区域，再添加 OCR 映射。")
                return

        self.mappings.append(mapping)
        self.refresh_mapping_list()
        QMessageBox.information(self, "OCR 映射已添加",
            f"已添加 OCR 映射：{img_count} 张图片 → "
            f"{len(mapping['target_cells'])} 个目标单元格\n"
            f"模式: {mode} | 预处理: {mapping.get('preprocess', 'none')}")

    # ==================== 映射列表刷新 ====================
    def refresh_mapping_list(self):
        self.mapping_list.clear()
        shown = 0
        for m in self._dedupe_mappings(self.mappings):
            if m.get('target_sheet') != self.current_sheet_name:
                continue
            shown += 1
            desc = f"{shown}. "
            if m['type'] == 'data':
                trans = m.get('transform', 'none')
                trans_str = f" [转换:{trans}]" if trans != 'none' else ""
                if m.get('clear_target'):
                    trans_str += " [先清空目标区]"
                desc += f"数据: {m['target_range']} <- {m['source_sheet']}!{m['source_range']}{trans_str}"
            elif m['type'] == 'image':
                align = m.get('alignment', 'left')
                align_cn = {'left': '左', 'center': '中', 'right': '右'}.get(align, align)
                desc += f"图片: 锚点{m['anchor_cell']} (列宽{m.get('col_width_chars','?')} 行高{m.get('row_height_pts','?')} 旋转{m.get('rotation',0)}° 缩放{m.get('width_scale',1.0)}x{m.get('height_scale',1.0)} {align_cn}对齐)"
            elif m['type'] == 'archive_shift_right':
                src_col = m.get('source_col', '?')
                desc += f"归档: {m['block_range']} <- {src_col}列 新表头\"{m['new_headers']}\""
            elif m['type'] == 'jmp':
                headers_str = ','.join(m['header_cols'])
                desc += f"JMP: 锚点{m['anchor_cell']} <- {m['source_sheet']}!{m['source_range']} (表头:{headers_str} 拼接:{m['merge_columns']})"
            elif m['type'] == 'ocr':
                img_count = len(m.get('image_list', []))
                mode = m.get('mode', '?')
                labels_str = ','.join(m.get('labels', []))
                mode_desc = {'first_number': '单值', 'labeled': f'多值({labels_str})',
                            'all_numbers': '全部数值', 'custom': '自定义'}.get(mode, mode)
                desc += f"OCR: {mode_desc} <- {img_count}张切片 (预处理:{m.get('preprocess','none')})"
            item = QListWidgetItem(desc)
            item.setData(Qt.UserRole, next(i for i, x in enumerate(self.mappings) if x is m))
            self.mapping_list.addItem(item)

    # ==================== 映射列表右键操作 ====================
    def show_mapping_context_menu(self, pos):
        item = self.mapping_list.itemAt(pos)
        if item is None:
            return
        mapping_idx = item.data(Qt.UserRole)
        menu = QMenu()
        menu.addAction(QAction("删除该映射", self, triggered=lambda: self.delete_mapping(mapping_idx)))
        menu.addSeparator()
        menu.addAction(QAction(f"清空本Sheet“{self.current_sheet_name}”全部映射", self,
                               triggered=self.clear_sheet_mappings))
        menu.addAction(QAction("清空全部映射", self, triggered=self.clear_all_mappings))
        menu.exec_(self.mapping_list.viewport().mapToGlobal(pos))

    def delete_mapping(self, mapping_idx):
        """按面板条目对应的索引删除映射"""
        if mapping_idx is not None and 0 <= mapping_idx < len(self.mappings):
            del self.mappings[mapping_idx]
            self.refresh_mapping_list()

    def clear_sheet_mappings(self):
        count = sum(1 for m in self.mappings if m.get('target_sheet') == self.current_sheet_name)
        if count == 0:
            return
        reply = QMessageBox.question(self, "确认清空",
            f"确定删除当前Sheet“{self.current_sheet_name}”的全部 {count} 条映射？",
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        self.mappings = [m for m in self.mappings if m.get('target_sheet') != self.current_sheet_name]
        self.refresh_mapping_list()

    def clear_all_mappings(self):
        if not self.mappings:
            return
        reply = QMessageBox.question(self, "确认清空",
            f"确定删除全部 {len(self.mappings)} 条映射？",
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        self.mappings = []
        self.refresh_mapping_list()

    # ==================== 确认映射 ====================
    def confirm_mapping(self):
        if not self.mappings:
            QMessageBox.information(self, "提示", "当前没有任何映射，请先添加映射")
            return
        current_mappings = [m for m in self.mappings if m.get('target_sheet') == self.current_sheet_name]
        if not current_mappings:
            QMessageBox.information(self, "提示", f"当前Sheet“{self.current_sheet_name}”没有映射")
        else:
            QMessageBox.information(self, "确认", f"已确认当前Sheet“{self.current_sheet_name}”的 {len(current_mappings)} 条映射，待输出报告时统一执行。")

    # ==================== 输出报告 ====================
    def _apply_cell_edits(self, ws, sheet_name):
        """把单独的单元格更新（右键修改/版本号查找）重放到工作表。
        必须在归档移位之后执行：归档会以磁盘数据覆盖区域，重放可保证
        用户手动更新落在最终版式上；若编辑单元格位于归档区块内，
        则随区块右移一格（保持编辑与数据一起移动）。"""
        edits = [e for e in self.cell_edits if e[0] == sheet_name]
        if not edits:
            return
        archive_blocks = [
            m['block_range'] for m in self._dedupe_mappings(self.mappings)
            if m.get('target_sheet') == sheet_name
            and m.get('type') == 'archive_shift_right']
        for entry in edits:
            row, col, value = entry[1], entry[2], entry[3]
            for min_row, min_col, max_row, max_col in archive_blocks:
                if min_row <= row <= max_row and min_col <= col <= max_col:
                    col += 1
            ws.cell(row=row, column=col).value = value

    @staticmethod
    def _first_date_header(ws):
        """查找表中第一个 Date 表头（不同制程位置可能略有变化）"""
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == 'date':
                    return cell
        return None

    @staticmethod
    def _fill_report_date(ws):
        """把 Summary 中 Date 表头右侧的单元格填充为报告当天日期"""
        header = MainWindow._first_date_header(ws)
        if header is None:
            return
        ws.cell(row=header.row, column=header.column + 1).value = datetime.datetime.now()

    # Build Phase / Configuration / Event 从文件夹名称解析
    BUILD_INFO_PATTERNS = {
        # 用"非字母数字"作边界（下划线不算边界，文件名/文件夹名都能匹配）
        'build_phase': re.compile(r'(?<![A-Za-z0-9])C\d+\.\d+(?![A-Za-z0-9])'),
        'configuration': re.compile(r'(?<![A-Za-z0-9])C\d{4}(?![A-Za-z0-9])'),
        'event': re.compile(r'(?<![A-Za-z0-9])(MBO|PBO)(?![A-Za-z0-9])', re.I),
    }

    @classmethod
    def _parse_build_info_from_folder(cls, folder_name):
        """从文件夹名称解析 Build Phase / Configuration / Event"""
        info = {}
        for key, pattern in cls.BUILD_INFO_PATTERNS.items():
            m = pattern.search(folder_name)
            if m:
                info[key] = m.group(0).upper() if key == 'event' else m.group(0)
        return info

    @classmethod
    def _parse_build_info_from_folders(cls, folder_names):
        """按顺序在多个文件夹名称中查找，每项取第一个命中"""
        info = {}
        for folder_name in folder_names:
            parsed = cls._parse_build_info_from_folder(folder_name)
            for key in ('build_phase', 'configuration', 'event'):
                if key not in info and parsed.get(key):
                    info[key] = parsed[key]
        return info

    @staticmethod
    def _find_build_info_value_cells(ws):
        """定位 Build Phase/Configuration/Event 表头下一格的值单元格"""
        targets = {}
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                norm = re.sub(r'\s+', ' ', cell.value).strip().lower()
                if norm == 'build phase':
                    targets['build_phase'] = (cell.row + 1, cell.column)
                elif norm == 'configuration':
                    targets['configuration'] = (cell.row + 1, cell.column)
                elif norm == 'event ( mbo /pbo )':
                    targets['event'] = (cell.row + 1, cell.column)
        return targets

    @staticmethod
    def _fill_build_info_from_folders(ws, folder_names):
        """从文件夹名称（按顺序查找）填充 Build Phase/Configuration/Event。
        返回未获取到的字段名列表（跳过更新，由用户手动填写）。"""
        info = MainWindow._parse_build_info_from_folders(folder_names)
        value_cells = MainWindow._find_build_info_value_cells(ws)
        failures = []
        for key, (row, col) in value_cells.items():
            value = info.get(key)
            if value:
                ws.cell(row=row, column=col).value = value
            else:
                failures.append(key)
        return failures

    @classmethod
    def _build_output_filename(cls, ws, template_path):
        """沿用模板文件名格式，把 Configuration/Event 段更新为 Summary
        中当前的值（文件名不含 Build Phase）；其余部分保留，
        由用户在保存对话框里自行修改。"""
        base_name = os.path.basename(template_path or '') or '报告'
        if base_name.lower().endswith('.xlsx'):
            base_name = base_name[:-5]
        elif '.' in base_name:
            base_name = base_name.rsplit('.', 1)[0]

        values = {}
        for key, (row, col) in cls._find_build_info_value_cells(ws).items():
            v = ws.cell(row=row, column=col).value
            if v is not None:
                values[key] = str(v).strip()

        name = base_name
        if values.get('configuration'):
            name = cls.BUILD_INFO_PATTERNS['configuration'].sub(
                values['configuration'], name, count=1)
        if values.get('event'):
            name = cls.BUILD_INFO_PATTERNS['event'].sub(values['event'], name, count=1)
        return name + '.xlsx'

    def output_report(self):
        if not self.template_wb:
            return
        checked_sheets = [self.sheet_list.item(i).text() for i in range(self.sheet_list.count())
                          if self.sheet_list.item(i).checkState() == Qt.Checked]
        if not checked_sheets:
            QMessageBox.warning(self, "警告", "请至少勾选一个需要输出的Sheet")
            return

        data_template_wb = openpyxl.load_workbook(self.template_path, data_only=True)
        data_source_wb = openpyxl.load_workbook(self.source_path, data_only=True) if self.source_path else None

        failed_mappings = []
        self._fill_warnings = []
        build_info_failures = []
        # 查找范围：从 IPQC 数据所在文件夹开始，找不到再往上一级
        search_folder_names = []
        base_dir = (os.path.dirname(self.source_path)
                    if self.source_path else
                    (os.path.dirname(self.template_path) if self.template_path else ''))
        if base_dir:
            search_folder_names.append(os.path.basename(base_dir) or base_dir)
            parent_dir = os.path.dirname(base_dir)
            if parent_dir:
                search_folder_names.append(os.path.basename(parent_dir) or parent_dir)

        # 预计算总映射数，用于进度条
        total_mappings = sum(
            len(self._dedupe_mappings(
                [m for m in self.mappings if m.get('target_sheet') == sheet_name]))
            for sheet_name in checked_sheets
        )
        prog = self._make_progress("正在处理中，请稍等！", total_mappings)
        done = 0

        for sheet_name in checked_sheets:
            ws = self.template_wb[sheet_name]
            data_ws = data_template_wb[sheet_name]
            # Summary：Date 表头右侧自动填充报告当天日期
            if 'summary' in sheet_name.lower():
                self._fill_report_date(ws)
                for key in self._fill_build_info_from_folders(ws, search_folder_names):
                    build_info_failures.append(f"{sheet_name}/{key}")
            sheet_mappings = self._dedupe_mappings(
                [m for m in self.mappings if m.get('target_sheet') == sheet_name])

            # 输出前防御：归档区块重叠会互相覆盖，直接报错并跳过，避免生成错误报告
            # （配置文件中加载进来的旧映射可能绕过添加时的冲突检查）
            archive_mappings = [m for m in sheet_mappings if m.get('type') == 'archive_shift_right']
            conflicted_ids = set()
            for i in range(len(archive_mappings)):
                for j in range(i + 1, len(archive_mappings)):
                    if self._archive_blocks_overlap(archive_mappings[i]['block_range'],
                                                    archive_mappings[j]['block_range']):
                        conflicted_ids.add(id(archive_mappings[i]))
                        conflicted_ids.add(id(archive_mappings[j]))
                        failed_mappings.append(
                            f"{sheet_name}/archive_shift_right: 归档区块重叠，已跳过："
                            f"{self._block_range_str(archive_mappings[i]['block_range'])} 与 "
                            f"{self._block_range_str(archive_mappings[j]['block_range'])}")

            # 执行顺序：归档 → 数据/图片 → 单独单元格更新 → JMP（最后）
            archive_mappings = [m for m in sheet_mappings
                                if m.get('type') == 'archive_shift_right']
            middle_mappings = [m for m in sheet_mappings
                               if m.get('type') not in ('archive_shift_right', 'jmp')]
            jmp_mappings = [m for m in sheet_mappings if m.get('type') == 'jmp']

            # PBO→MBO 整区更新：先按区域一次性清空数值+图片（数据映射勾选“先清空目标区”，
            # 图片批量映射自动整区清空），再执行填充，避免扩展区域内残留 PBO 旧数据/旧图片。
            # 同一区域只清一次，与映射添加顺序无关。
            cleared_regions = set()
            for m in middle_mappings:
                region = None
                if m.get('type') == 'data' and m.get('clear_target'):
                    region = m.get('target_range')
                elif m.get('type') == 'image' and m.get('clear_region'):
                    region = m.get('clear_region')
                if region:
                    key = (sheet_name, tuple(region))
                    if key not in cleared_regions:
                        cleared_regions.add(key)
                        self.clear_region(ws, region)

            def run_mapping(mapping):
                nonlocal done
                if id(mapping) in conflicted_ids:
                    return
                # 按映射自身的 source_sheet 解析数据源（此前误用目标Sheet名查找，
                # 数据源存在同名Sheet时会静默读错表）
                data_src_ws = self._resolve_data_src_ws(mapping, data_source_wb)
                # JMP 的源表在模板文件内，取其数值版（data_only），不能用目标Sheet代替
                jmp_src_ws = None
                if mapping.get('type') == 'jmp':
                    src_sheet = mapping.get('source_sheet')
                    if src_sheet and src_sheet in data_template_wb.sheetnames:
                        jmp_src_ws = data_template_wb[src_sheet]
                try:
                    self.execute_mapping(ws, mapping, data_ws, data_src_ws, jmp_src_ws)
                except Exception as e:
                    failed_mappings.append(f"{sheet_name}/{mapping.get('type')}: {e}")
                done += 1
                self._update_progress(prog, done, total_mappings)

            for mapping in archive_mappings + middle_mappings:
                run_mapping(mapping)
            # 归档与数据/图片之后，重放该表的单独单元格更新（右键修改/版本号）
            self._apply_cell_edits(ws, sheet_name)
            # JMP 最后执行
            for mapping in jmp_mappings:
                run_mapping(mapping)

        self._close_progress(prog)

        if build_info_failures:
            QMessageBox.warning(
                self, "部分字段未自动更新",
                "以下字段无法从文件夹名称获取，已跳过更新"
                f"（已查找：{' → '.join(search_folder_names) or '无'}）：\n"
                + "\n".join(build_info_failures)
                + "\n\n请在报告中手动修改这些单元格。")

        if self._fill_warnings:
            shown = self._fill_warnings[:20]
            more = len(self._fill_warnings) - len(shown)
            msg = "处理过程中发现以下警告，请检查后再确认报告：\n\n" + "\n".join(shown)
            if more > 0:
                msg += f"\n…另有 {more} 处"
            QMessageBox.warning(self, "处理警告", msg)

        data_template_wb.close()
        if data_source_wb:
            data_source_wb.close()

        # 默认文件名：沿用模板格式，Configuration/Event 用 Summary 当前值
        summary_sheet = next((s for s in checked_sheets if 'summary' in s.lower()), None)
        default_name = ''
        if summary_sheet:
            default_name = self._build_output_filename(
                self.template_wb[summary_sheet], self.template_path)
        default_dir = os.path.dirname(self.template_path) if self.template_path else ''
        default_path = os.path.join(default_dir, default_name) if default_dir else default_name
        save_path, _ = QFileDialog.getSaveFileName(
            self, "保存报告", default_path, "Excel文件 (*.xlsx)")
        if save_path:
            save_prog = self._make_busy_progress("正在保存报告，请稍等…")
            try:
                self._refresh_image_refs()
                self.template_wb.save(save_path)
            except Exception as e:
                self._close_progress(save_prog)
                QMessageBox.critical(self, "保存失败", f"保存报告时出错：{e}")
                return
            self._close_progress(save_prog)

            self._image_streams.clear()

            if failed_mappings:
                QMessageBox.warning(self, "部分映射失败",
                    "以下映射未能成功执行：\n" + "\n".join(failed_mappings) +
                    "\n\n报告已保存，请手动检查这些区域。")
            else:
                QMessageBox.information(self, "完成", "报告已保存")


    # ==================== 配置保存 ====================
    def save_config(self):
        if not self.template_path:
            QMessageBox.warning(self, "提示", "请先打开模板文件")
            return
        clean_mappings = []
        for m in self.mappings:
            m_copy = m.copy()
            # 配置只记录映射路径（从哪里→到哪里），不保存内容本身
            m_copy.pop('image_bytes', None)
            clean_mappings.append(m_copy)

        config = {'template_file': self.template_path, 'mappings': clean_mappings}
        default_dir = os.path.dirname(self.template_path) or ''
        default_name = os.path.basename(self.template_path).rsplit('.', 1)[0] + '_config.json'
        default_path = os.path.join(default_dir, default_name) if default_dir else default_name
        config_path, _ = QFileDialog.getSaveFileName(
            self, "保存配置", default_path, "配置文件 (*.json)")
        if not config_path:
            return
        if not config_path.lower().endswith('.json'):
            config_path += '.json'
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            QMessageBox.information(self, "完成", f"配置已保存至 {config_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"配置保存失败：{e}")

    def import_config(self):
        if not self.template_wb:
            QMessageBox.warning(self, "提示", "请先打开报告文件")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "导入我的配置", "", "配置文件 (*.json);;所有文件 (*)")
        if not path:
            return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            mappings = config.get('mappings', [])
            if not isinstance(mappings, list):
                raise ValueError("配置中 mappings 格式不正确")
            cfg_template = config.get('template_file')
            if (cfg_template and self.template_path
                    and os.path.basename(cfg_template) != os.path.basename(self.template_path)):
                reply = QMessageBox.question(
                    self, "确认导入",
                    f"该配置来自模板：\n{os.path.basename(cfg_template)}\n"
                    f"与当前报告：\n{os.path.basename(self.template_path)}\n"
                    "不一致，仍要导入吗？",
                    QMessageBox.Yes | QMessageBox.No)
                if reply != QMessageBox.Yes:
                    return
            self.mappings = normalize_mappings(mappings)
            self.refresh_mapping_list()
            QMessageBox.information(self, "完成", f"已导入 {len(mappings)} 条映射")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"配置导入失败：{e}")


# ==================== 用户ID验证 ====================
class UserIdDialog(QDialog):
    """欢迎界面之后、进入主界面之前的用户ID验证窗口"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("用户ID验证")
        self.setFixedWidth(360)
        self.user_id = None

        layout = QVBoxLayout(self)
        tip = QLabel("请输入自己的工号：")
        layout.addWidget(tip)

        self.edit = QLineEdit()
        self.edit.setPlaceholderText("请输入工号")
        self.edit.setMaxLength(32)
        self.edit.returnPressed.connect(self.try_login)
        layout.addWidget(self.edit)

        self.error_label = QLabel("")
        self.error_label.setStyleSheet("color: #b00020; font-weight: bold;")
        self.error_label.setVisible(False)
        layout.addWidget(self.error_label)

        btn_layout = QHBoxLayout()
        self.btn_ok = QPushButton("确定")
        self.btn_ok.clicked.connect(self.try_login)
        self.btn_cancel = QPushButton("退出")
        self.btn_cancel.clicked.connect(self.reject)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_ok)
        btn_layout.addWidget(self.btn_cancel)
        layout.addLayout(btn_layout)

        self.edit.setFocus()

    def try_login(self):
        user_id = self.edit.text().strip().upper()
        if not user_id:
            self._show_unauthorized()
            return
        if user_id == ADMIN_USER_ID:
            self._admin_flow()
            return
        if user_id in load_authorized_ids():
            self.user_id = user_id
            self.accept()
            return
        self._show_unauthorized()

    def _show_unauthorized(self):
        self.error_label.setText("您当前ID未授权！")
        self.error_label.setVisible(True)
        self.edit.selectAll()
        self.edit.setFocus()

    def _admin_flow(self):
        """管理员流程：授权密码 → ID授权窗口 → 进入主界面"""
        password, ok = QInputDialog.getText(
            self, "管理员授权", "请输入授权密码：", QLineEdit.Password)
        if not ok:
            return
        if password != ADMIN_AUTH_PASSWORD:
            QMessageBox.warning(self, "密码错误", "授权密码错误！")
            return
        dlg = AuthorizeIdDialog(self)
        dlg.exec_()
        new_ids = dlg.get_new_ids()
        if new_ids:
            current = load_authorized_ids()
            combined = list(dict.fromkeys(current + new_ids))
            saved = save_authorized_ids(combined)
            msg = f"已授权 {len(new_ids)} 个工号：\n" + ", ".join(new_ids)
            if not saved:
                msg += "\n（授权文件保存失败，本次运行内有效）"
            QMessageBox.information(self, "授权完成", msg)
        self.user_id = ADMIN_USER_ID
        self.accept()

    def get_user_id(self):
        return self.user_id


class AuthorizeIdDialog(QDialog):
    """管理员授权窗口：批量输入需要授权的工号"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("ID授权")
        self.setMinimumSize(460, 300)

        layout = QVBoxLayout(self)
        tip = QLabel(
            "请输入需要授权的工号，多个工号可用空格、逗号、分号、斜杠等隔开：")
        tip.setWordWrap(True)
        layout.addWidget(tip)

        self.text = QPlainTextEdit()
        self.text.setPlaceholderText("例如：G1655895 G1659304; G1234567, G7654321")
        layout.addWidget(self.text, 1)

        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("确定")
        btn_ok.clicked.connect(self.accept)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)

    def get_new_ids(self):
        return parse_id_input(self.text.toPlainText())



SPLASH_MIN_MS = 1000  # 启动画面最短展示时间
FADE_MS = 300         # 淡出过渡时长（Apple 风格，平滑衔接主窗口）
_ANIMS = set()        # 持有运行中的动画引用，防止被提前回收


def create_splash(target_size=None):
    """按 splash.png 生成启动画面，尺寸与主窗口一致（cover 裁切填满）。"""
    pixmap = QPixmap(resource_path('splash.png'))
    if pixmap.isNull():
        # 备用：纯白背景 + 文字
        w, h = target_size or window_target_size()
        pixmap = QPixmap(w, h)
        pixmap.fill(Qt.white)
        painter = QPainter(pixmap)
        painter.setFont(QFont('Arial', 20))
        painter.drawText(pixmap.rect(), Qt.AlignCenter, "自动报告工具")
        painter.end()
        return QSplashScreen(pixmap)
    w, h = target_size or window_target_size()
    img = pixmap.toImage()
    scale = max(w / img.width(), h / img.height())
    sw, sh = int(img.width() * scale), int(img.height() * scale)
    img = img.scaled(sw, sh, Qt.IgnoreAspectRatio, Qt.SmoothTransformation)
    x = max(0, (sw - w) // 2)
    y = max(0, (sh - h) // 2)
    img = img.copy(x, y, w, h)
    return QSplashScreen(QPixmap.fromImage(img))


def _crossfade(splash, window, duration=FADE_MS, on_finished=None):
    """启动画面平滑淡出，露出下方已就绪的主窗口，避免生硬切换。"""
    fade_out = QPropertyAnimation(splash, b"windowOpacity")
    fade_out.setDuration(duration)
    fade_out.setStartValue(1.0)
    fade_out.setEndValue(0.0)
    fade_out.setEasingCurve(QEasingCurve.InOutQuad)
    _ANIMS.add(fade_out)

    def _finish():
        splash.hide()
        splash.finish(window)
        _ANIMS.discard(fade_out)
        if on_finished:
            on_finished()

    fade_out.finished.connect(_finish)
    fade_out.start()


if __name__ == '__main__':
    # PyInstaller 冻结环境下 multiprocessing spawn 子进程需要此调用
    multiprocessing.freeze_support()
    # Windows 高分屏适配（必须在 QApplication 创建前设置）
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    app.setApplicationName("自动M/PBO报告制作软件")
    # Apple 风格全局样式（主窗口 + 所有对话框统一）
    app.setStyleSheet(APPLE_QSS)
    icon = QIcon(resource_path('app_icon.ico'))
    if not icon.isNull():
        app.setWindowIcon(icon)

    # 启动画面：与主窗口同尺寸（cover 裁切），最短展示后淡出
    splash_start = time.monotonic()
    splash = create_splash()
    if splash is not None:
        splash.show()
        splash.showMessage(
            "正在启动 自动M/PBO报告制作软件 ...",
            Qt.AlignBottom | Qt.AlignCenter,
            Qt.white,
        )
        app.processEvents()

    window = MainWindow()
    if not icon.isNull():
        window.setWindowIcon(icon)
    window.show()

    def start_login():
        # 用户工号验证：以主界面为背景，通过后才可使用软件功能
        login = UserIdDialog(window)
        if login.exec_() != QDialog.Accepted:
            sys.exit(0)

    if splash is not None:
        elapsed_ms = int((time.monotonic() - splash_start) * 1000)
        remaining_ms = max(0, SPLASH_MIN_MS - elapsed_ms)
        QTimer.singleShot(remaining_ms, lambda: _crossfade(splash, window, on_finished=start_login))
    else:
        start_login()

    sys.exit(app.exec_())
